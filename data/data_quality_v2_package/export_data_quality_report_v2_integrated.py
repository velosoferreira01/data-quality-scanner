# -*- coding: utf-8 -*-
"""
export_data_quality_report_v2_integrated.py

Relatório premium executivo de Data Quality — aderência percentual e revisão executiva com GenAI/Ollama:
- usa um único run_id efetivo
- gera HTML com gráficos embutidos
- gera Excel com abas
- exporta radar, histórico e arquivos CSV auxiliares
"""

import argparse
import base64
import json
import os
import re
import sys
import time
import unicodedata
from datetime import datetime
from io import BytesIO
from pathlib import Path
from urllib import error as urllib_error
from urllib import request as urllib_request

import duckdb
import matplotlib.pyplot as plt
import pandas as pd


# =========================================================
# Tradução da camada de apresentação (PT-BR)
# =========================================================
# Os nomes técnicos continuam em inglês no DuckDB e nos CSVs auxiliares.
# A tradução abaixo é aplicada somente aos relatórios HTML e Excel.
COLUMN_LABELS_PT_BR = {
    # Execução e origem
    "run_id": "ID da execução",
    "scanned_at": "Data e hora da análise",
    "started_at": "Início da execução",
    "finished_at": "Fim da execução",
    "created_at": "Data de criação",
    "duration_seconds": "Duração em segundos",
    "source_type": "Tipo de origem",
    "source_ref": "Referência da origem",
    "source_name": "Nome da origem",
    "source": "Origem",
    "file_name": "Nome do arquivo",
    "schema_name": "Nome do esquema",
    "table_name": "Nome da tabela",
    "object_name": "Nome do objeto",

    # Dataset / conjunto de dados
    "dataset": "Conjunto de dados",
    "datasets": "Conjuntos de dados",
    "dataset_name": "Nome do conjunto de dados",
    "dataset_label": "Conjunto de dados",
    "row_count": "Quantidade de registros",
    "column_count": "Quantidade de colunas",
    "record_type_count": "Quantidade de tipos de registro",
    "records_by_type": "Quantidade de registros por tipo",

    # Regras, validações e colunas
    "validation_code": "Código da validação",
    "validation_name": "Nome da validação",
    "rule_code": "Código da regra",
    "rule_name": "Nome da regra",
    "rule_type": "Tipo da regra",
    "column_name": "Nome da coluna",
    "related_columns": "Colunas relacionadas",
    "dtype": "Tipo de dado",
    "metric_name": "Nome da métrica",
    "metric_value": "Valor da métrica",

    # Quantidades e taxas
    "total_count": "Total de registros",
    "evaluated_count": "Registros avaliados",
    "valid_count": "Registros válidos",
    "invalid_count": "Registros inválidos",
    "null_count": "Registros nulos",
    "error_count": "Quantidade de erros",
    "passed_count": "Registros aprovados",
    "failed_count": "Registros reprovados",
    "valid_rate": "Taxa de validade",
    "invalid_rate": "Taxa de invalidade",
    "null_rate": "Taxa de valores nulos",
    "error_rate": "Taxa de erro",
    "distinct_ratio": "Proporção de valores distintos",
    "coverage_rate": "Taxa de cobertura",
    "coverage": "Cobertura",
    "violations": "Violações",

    # Pontuação e classificação
    "score": "Pontuação",
    "score_base": "Pontuação base",
    "score_final": "Pontuação final",
    "score_bacen": "Pontuação BACEN",
    "avg_score": "Pontuação média",
    "average_score": "Pontuação média",
    "minimum_score": "Pontuação mínima",
    "minimum_score_observed": "Menor pontuação observada",
    "maximum_score_observed": "Maior pontuação observada",
    "raw_score": "Pontuação bruta",
    "weighted_score": "Pontuação ponderada",
    "weight": "Peso",
    "classification": "Classificação",
    "classification_bacen": "Classificação BACEN",
    "classificacao": "Classificação",
    "status": "Status",
    "severity": "Severidade",
    "priority": "Prioridade",
    "prioridade": "Prioridade",

    # Dimensões BACEN
    "dimension_code": "Código da dimensão",
    "dimension_name": "Nome da dimensão",
    "description": "Descrição",
    "evaluation_type": "Tipo de avaliação",
    "applicable": "Aplicável",
    "evidence": "Evidências",
    "evaluated_datasets": "Conjuntos de dados avaliados",
    "total_datasets": "Total de conjuntos de dados",

    # Evidências, referências e recomendações
    "reference_source": "Fonte de referência",
    "details": "Detalhes",
    "detail": "Detalhe",
    "issue": "Problema identificado",
    "issues": "Problemas identificados",
    "recommendation": "Recomendação",
    "recomendacao": "Recomendação",
    "sample_value": "Valor de exemplo",
    "masked_value": "Valor mascarado",
    "row_number": "Número da linha",

    # KPIs executivos
    "score_medio": "Pontuação média",
    "melhor_dataset": "Melhor conjunto de dados",
    "melhor_score": "Melhor pontuação",
    "pior_dataset": "Pior conjunto de dados",
    "pior_score": "Pior pontuação",
    "criticos": "Itens críticos",
    "atencao": "Itens em atenção",
    "aceitaveis": "Itens aceitáveis",
    "satisfatorios": "Itens satisfatórios",
    "bons": "Itens bons",
    "excelentes": "Itens excelentes",
    "aderencia_percentual": "Aderência de qualidade",
    "melhor_aderencia": "Melhor aderência",
    "pior_aderencia": "Pior aderência",

    # Dimensões técnicas legadas
    "dim_acessibilidade": "Acessibilidade",
    "dim_acuracia": "Acurácia",
    "dim_adaptabilidade": "Adaptabilidade",
    "dim_clareza": "Clareza",
    "dim_comparabilidade": "Comparabilidade",
    "dim_completude": "Completude",
    "dim_confiabilidade": "Confiabilidade",
    "dim_consistencia": "Consistência",
    "dim_integridade": "Integridade",
    "dim_integridade_ref": "Integridade referencial",
    "dim_rastreabilidade": "Rastreabilidade",
    "dim_relevancia": "Relevância",
    "dim_tempestividade": "Tempestividade",
    "dim_unicidade": "Unicidade",
    "dim_validade": "Validade",
    "dim_freshness": "Atualidade",
}

VALUE_LABELS_PT_BR = {
    "PASS": "Aprovado",
    "PASSED": "Aprovado",
    "FAIL": "Reprovado",
    "FAILED": "Reprovado",
    "SUCCESS": "Sucesso",
    "ERROR": "Erro",
    "WARNING": "Atenção",
    "CRITICAL": "Crítico",
    "VALID": "Válido",
    "INVALID": "Inválido",
    "VALIDO": "Válido",
    "INVALIDO": "Inválido",
    "COMPLIANT": "Conforme",
    "NON_COMPLIANT": "Não conforme",
    "CONFORME": "Conforme",
    "NAO_CONFORME": "Não conforme",
    "NÃO CONFORME": "Não conforme",
    "APPLICABLE": "Aplicável",
    "NOT_APPLICABLE": "Não aplicável",
    "TRUE": "Sim",
    "FALSE": "Não",
    "YES": "Sim",
    "NO": "Não",
    "HIGH": "Alta",
    "MEDIUM": "Média",
    "LOW": "Baixa",
    "ADEQUAÇÃO": "Satisfatório",
    "ADEQUACAO": "Satisfatório",
    "GOVERNANÇA": "Governança",
    "GOVERNANCA": "Governança",
    "HIBRIDO": "Híbrido",
    "HÍBRIDO": "Híbrido",
    "AUTOMATIZADO": "Automatizado",
    "AUTOMATED": "Automatizado",
    "AUTOMATIC": "Automatizado",
    "HYBRID": "Híbrido",
    "GOVERNANCE": "Governança",
    "NÃO AVALIADO": "Não avaliado",
    "NAO AVALIADO": "Não avaliado",
}


def _unique_labels(labels):
    """Evita títulos duplicados após a tradução."""
    counts = {}
    unique = []
    for label in labels:
        counts[label] = counts.get(label, 0) + 1
        if counts[label] == 1:
            unique.append(label)
        else:
            unique.append(f"{label} ({counts[label]})")
    return unique


EVALUATION_TYPE_LABELS_PT_BR = {
    "automated": "Automatizado",
    "automatic": "Automatizado",
    "automatizado": "Automatizado",
    "automático": "Automatizado",
    "automatico": "Automatizado",
    "hybrid": "Híbrido",
    "híbrido": "Híbrido",
    "hibrido": "Híbrido",
    "governance": "Governança",
    "governance-based": "Governança",
    "governance based": "Governança",
    "governança": "Governança",
    "governanca": "Governança",
    "manual": "Manual",
}


def translate_evaluation_type(value):
    """
    Traduz tipos de avaliação simples ou combinados.

    Exemplos:
    - automated -> Automatizado
    - hybrid -> Híbrido
    - governance | hybrid -> Governança | Híbrido
    """
    if not isinstance(value, str):
        return value

    raw = value.strip()
    if not raw:
        return value

    normalized = raw.replace(";", "|").replace(",", "|").replace("/", "|")
    parts = [part.strip() for part in normalized.split("|") if part.strip()]
    if not parts:
        return value

    translated_parts = []
    for part in parts:
        translated = EVALUATION_TYPE_LABELS_PT_BR.get(part.lower())
        if translated is None:
            return value
        if translated not in translated_parts:
            translated_parts.append(translated)

    return " | ".join(translated_parts)


def _translate_cell_value(value):
    """Traduz somente valores textuais conhecidos, preservando números."""
    if isinstance(value, bool):
        return "Sim" if value else "Não"

    if not isinstance(value, str):
        return value

    evaluation_type = translate_evaluation_type(value)
    if evaluation_type != value:
        return evaluation_type

    stripped = value.strip()
    translated = VALUE_LABELS_PT_BR.get(stripped.upper())
    return translated if translated is not None else value


def prepare_dataframe_for_report(df):
    """
    Cria uma cópia do DataFrame para apresentação em HTML/Excel.

    A estrutura original utilizada pelo DuckDB e pelo pipeline não é alterada.
    """
    if df is None:
        return pd.DataFrame()

    show = df.copy()

    translated_columns = [
        COLUMN_LABELS_PT_BR.get(str(column), str(column))
        for column in show.columns
    ]
    show.columns = _unique_labels(translated_columns)

    for column in show.columns:
        if pd.api.types.is_object_dtype(show[column]) or pd.api.types.is_bool_dtype(show[column]):
            show[column] = show[column].map(_translate_cell_value)

    return show



# =========================================================
# Perfis e formatação profissional das tabelas
# =========================================================
TABLE_PROFILES = {
    "summary": [
        "dataset_label", "row_count", "column_count",
        "record_type_count", "records_by_type", "score_bacen",
        "score", "score_final", "classification_bacen",
        "classification", "status",
    ],
    "dimensions": [
        "dataset_label", "score_bacen", "dim_acessibilidade",
        "dim_acuracia", "dim_adaptabilidade", "dim_clareza",
        "dim_comparabilidade", "dim_completude", "dim_confiabilidade",
        "dim_consistencia", "dim_integridade", "dim_rastreabilidade",
        "dim_relevancia", "dim_tempestividade", "dim_unicidade",
        "dim_validade", "dim_integridade_ref", "dim_freshness",
    ],
    "bacen_catalog": [
        "dimension_name", "description", "average_score",
        "minimum_score_observed", "maximum_score_observed",
        "evaluated_datasets", "total_datasets", "weight",
        "evaluation_type", "status", "evidence",
    ],
    "bacen_detail": [
        "dataset_label", "dimension_name", "weight", "minimum_score",
        "evaluation_type", "raw_score", "weighted_score", "status",
        "applicable", "evidence",
    ],
    "semantic": [
        "object_name", "validation_name", "column_name", "related_columns",
        "total_count", "evaluated_count", "valid_count", "invalid_count",
        "null_count", "valid_rate", "status", "reference_source", "details",
    ],
    "semantic_invalid": [
        "object_name", "validation_name", "column_name", "row_number",
        "masked_value", "sample_value", "details",
    ],
    "attention": [
        "dataset_label", "column_name", "dtype", "score", "classification",
        "null_rate", "distinct_ratio", "violations",
    ],
    "history": ["run_id", "avg_score"],
    "top_bottom": ["dataset_label", "score", "classification"],
    "technical": None,
}

PROFILE_COLUMN_LABELS = {
    "summary": {
        "_summary_score": "Aderência de qualidade",
        "_summary_classification": "Classificação",
    },
    "dimensions": {
        "score_bacen": "Índice BACEN",
    },
    "bacen_catalog": {
        "dimension_name": "Dimensão",
        "average_score": "Média",
        "minimum_score_observed": "Menor média avaliada",
        "maximum_score_observed": "Maior média avaliada",
        "evaluated_datasets": "Conjuntos avaliados",
        "total_datasets": "Total de conjuntos",
        "weight": "Peso",
        "evaluation_type": "Tipo de avaliação",
        "status": "Classificação",
        "evidence": "Evidências",
    },
    "bacen_detail": {
        "dataset_label": "Conjunto de dados",
        "dimension_name": "Dimensão",
        "weight": "Peso",
        "minimum_score": "Aderência mínima esperada",
        "evaluation_type": "Tipo de avaliação",
        "raw_score": "Aderência obtida",
        "weighted_score": "Contribuição ponderada",
        "status": "Classificação",
        "applicable": "Aplicável",
        "evidence": "Evidência",
    },
    "attention": {"score": "Aderência de qualidade"},
    "history": {"avg_score": "Aderência média"},
    "top_bottom": {"score": "Aderência de qualidade"},
}

PERCENTAGE_COLUMNS = {
    "valid_rate", "invalid_rate", "null_rate", "error_rate",
    "distinct_ratio", "coverage_rate", "weight",
}
SCORE_COLUMNS = {
    "score", "score_base", "score_final", "score_bacen", "avg_score",
    "average_score", "minimum_score", "minimum_score_observed",
    "maximum_score_observed", "raw_score", "weighted_score",
    "melhor_score", "pior_score", "score_medio",
}
COUNT_COLUMNS = {
    "row_count", "column_count", "total_count", "evaluated_count",
    "valid_count", "invalid_count", "null_count", "error_count",
    "passed_count", "failed_count", "violations", "evaluated_datasets",
    "total_datasets",
}
DATE_COLUMNS = {"scanned_at", "started_at", "finished_at", "created_at"}
LONG_TEXT_COLUMNS = {
    "details", "detail", "evidence", "description", "recommendation",
    "recomendacao", "related_columns", "issues", "issue",
    "column_name", "records_by_type", "masked_value", "sample_value",
}


COLUMN_LAYOUT_CLASSES = {
    "dataset_label": "cell-dataset-name",
    "object_name": "cell-object-name",
    "column_name": "cell-column-name",
    "row_number": "cell-row-number",
    "masked_value": "cell-masked-value",
    "sample_value": "cell-sample-value",
    "records_by_type": "cell-records-by-type",
    "dtype": "cell-data-type",
}


def _column_layout_class(original_column):
    """Retorna uma classe de layout para controlar largura e quebra de texto."""
    return COLUMN_LAYOUT_CLASSES.get(str(original_column), "")


def _profile_column_label(profile, column):
    # Na visão por dimensão, os títulos ficam limpos: somente o nome da dimensão.
    if profile == "dimensions" and str(column).startswith("dim_"):
        return COLUMN_LABELS_PT_BR.get(
            str(column),
            str(column).replace("dim_", "").replace("_", " ").title(),
        )

    profile_labels = PROFILE_COLUMN_LABELS.get(profile, {})
    return profile_labels.get(
        str(column),
        COLUMN_LABELS_PT_BR.get(str(column), str(column)),
    )


def _is_adherence_score_column(profile, column):
    column = str(column)
    if profile in {None, "technical", "semantic", "semantic_invalid"}:
        return False
    if profile == "summary":
        return column == "_summary_score"
    if profile == "dimensions":
        return column == "score_bacen" or column.startswith("dim_")
    if profile == "bacen_catalog":
        return column in {
            "average_score", "minimum_score_observed", "maximum_score_observed"
        }
    if profile == "bacen_detail":
        return column in {"minimum_score", "raw_score"}
    if profile in {"attention", "top_bottom"}:
        return column == "score"
    if profile == "history":
        return column == "avg_score"
    return False


def _is_weighted_contribution_column(profile, column):
    return profile == "bacen_detail" and str(column) == "weighted_score"


def _is_false_like(value):
    return safe_str(value).strip().lower() in {
        "false", "0", "não", "nao", "n", "no", "não aplicável", "nao aplicavel"
    }


def normalize_status_label(value):
    """Padroniza classificações e impede rótulos ambíguos no relatório."""
    raw = safe_str(value).strip()
    if not raw:
        return "Não avaliado"

    key = raw.upper()
    mapping = {
        "ADEQUAÇÃO": "Satisfatório",
        "ADEQUACAO": "Satisfatório",
        "ADEQUADO": "Satisfatório",
        "ADEQUADA": "Satisfatório",
        "SATISFATORIO": "Satisfatório",
        "SATISFATORIOS": "Satisfatório",
        "CONFORME": "Conforme",
        "BOM": "Bom",
        "EXCELENTE": "Excelente",
        "ATENÇÃO": "Atenção",
        "ATENCAO": "Atenção",
        "CRÍTICO": "Crítico",
        "CRITICO": "Crítico",
        "DEFICIENTE": "Crítico",
        "NÃO": "Não avaliado",
        "NAO": "Não avaliado",
        "NÃO AVALIADO": "Não avaliado",
        "NAO AVALIADO": "Não avaliado",
        "SEM EVIDÊNCIA": "Não avaliado",
        "SEM EVIDENCIA": "Não avaliado",
        "SEM CLASSIFICAÇÃO": "Não avaliado",
        "SEM CLASSIFICACAO": "Não avaliado",
    }
    return mapping.get(key, raw[:1].upper() + raw[1:])




def _is_missing_metric_value(value):
    """Identifica valores ausentes sem confundir zero válido com ausência."""
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except Exception:
        return False


def _normalize_identifier(value):
    raw = safe_str(value).strip().lower()
    normalized = unicodedata.normalize("NFKD", raw)
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    return normalized.replace("-", "_").replace(" ", "_")


def apply_na_to_unmeasured_dimensions(dim_df, bacen_detail_df):
    """
    Substitui por ausência apenas dimensões que realmente não foram avaliadas.

    Um zero continua sendo exibido quando a validação ocorreu e o resultado foi
    efetivamente zero. O valor vira N/A quando não existe mensuração, quando o
    item não é aplicável ou quando o status indica falta de evidência.
    """
    if dim_df is None or dim_df.empty:
        return pd.DataFrame() if dim_df is None else dim_df
    if bacen_detail_df is None or bacen_detail_df.empty:
        return dim_df.copy()
    if not {"dataset_label", "dimension_code"}.issubset(bacen_detail_df.columns):
        return dim_df.copy()

    output = dim_df.copy()
    detail = bacen_detail_df.copy()
    detail["_dataset_key"] = detail["dataset_label"].map(_normalize_identifier)
    detail["_dimension_key"] = detail["dimension_code"].map(_normalize_identifier)

    for (dataset_key, dimension_key), part in detail.groupby(
        ["_dataset_key", "_dimension_key"], dropna=False
    ):
        dimension_column = f"dim_{dimension_key}"
        if dimension_column not in output.columns:
            continue

        measured_values = []
        for _, row in part.iterrows():
            status = normalize_status_label(row.get("status"))
            applicable = row.get("applicable", True)
            raw_score = pd.to_numeric(
                pd.Series([row.get("raw_score")]), errors="coerce"
            ).iloc[0]

            measured = (
                not _is_false_like(applicable)
                and status not in {
                    "Não avaliado", "Não aplicável", "Sem classificação"
                }
                and not pd.isna(raw_score)
            )
            if measured:
                measured_values.append(float(raw_score))

        dataset_mask = output["dataset_label"].map(
            _normalize_identifier
        ) == dataset_key

        if measured_values:
            output.loc[dataset_mask, dimension_column] = sum(measured_values) / len(measured_values)
        else:
            output.loc[dataset_mask, dimension_column] = float("nan")

    return output


RECORD_TYPE_COLUMN_CANDIDATES = {
    "record_type",
    "tipo_registro",
    "tipo_de_registro",
    "tp_registro",
    "recordtype",
    "tipo_linha",
    "record_kind",
    "tipo_do_registro",
}


def _display_record_type(value):
    text_value = safe_str(value).strip()
    if not text_value:
        return "Sem identificação"
    upper = text_value.upper()
    if upper == "HEADER":
        return "Cabeçalho"
    if upper == "TRAILER":
        return "Rodapé"
    try:
        numeric = float(text_value)
        if numeric.is_integer():
            return f"Tipo {int(numeric)}"
    except Exception:
        pass
    return text_value[:1].upper() + text_value[1:]


def _record_type_sort_key(label):
    normalized = _normalize_identifier(label)
    if normalized == "cabecalho":
        return (0, 0, label)
    if normalized.startswith("tipo_"):
        try:
            return (1, int(normalized.split("_", 1)[1]), label)
        except Exception:
            return (1, 999, label)
    if normalized == "rodape":
        return (3, 0, label)
    return (2, 0, label)


def _match_input_file(input_dir, row):
    input_dir = Path(input_dir)
    candidates = [
        row.get("source_ref"), row.get("source"), row.get("file_name"),
        row.get("dataset_label"), row.get("dataset_name"), row.get("object_name"),
    ]

    for candidate in candidates:
        candidate_text = safe_str(candidate)
        if not candidate_text:
            continue
        candidate_path = Path(candidate_text)
        if candidate_path.exists() and candidate_path.is_file():
            return candidate_path

    if not input_dir.exists():
        return None

    files = [path for path in input_dir.iterdir() if path.is_file()]
    for candidate in candidates:
        candidate_text = safe_str(candidate).split("::", 1)[0]
        if not candidate_text:
            continue
        candidate_name = Path(candidate_text).name.lower()
        candidate_stem = Path(candidate_text).stem.lower()
        for path in files:
            if path.name.lower() == candidate_name or path.stem.lower() == candidate_stem:
                return path
    return None


def _detect_record_type_column(columns):
    normalized_to_original = {
        _normalize_identifier(column): column for column in columns
    }
    for candidate in RECORD_TYPE_COLUMN_CANDIDATES:
        if candidate in normalized_to_original:
            return normalized_to_original[candidate]
    return None


def _read_record_type_counts(path, fallback_row_count=None):
    """Lê somente a coluna de tipo, evitando carregar colunas desnecessárias."""
    suffix = path.suffix.lower()
    record_type_column = None
    counts = {}

    try:
        if suffix == ".csv":
            header = pd.read_csv(path, nrows=0)
            record_type_column = _detect_record_type_column(header.columns)
            if record_type_column:
                for chunk in pd.read_csv(
                    path,
                    usecols=[record_type_column],
                    dtype=str,
                    chunksize=100_000,
                    low_memory=False,
                ):
                    series = chunk[record_type_column].fillna("").map(_display_record_type)
                    for label, quantity in series.value_counts(dropna=False).items():
                        counts[label] = counts.get(label, 0) + int(quantity)
        elif suffix in {".xlsx", ".xls"}:
            header = pd.read_excel(path, nrows=0)
            record_type_column = _detect_record_type_column(header.columns)
            if record_type_column:
                frame = pd.read_excel(path, usecols=[record_type_column], dtype=str)
                series = frame[record_type_column].fillna("").map(_display_record_type)
                counts = {label: int(quantity) for label, quantity in series.value_counts().items()}
        elif suffix == ".parquet":
            header = pd.read_parquet(path).head(0)
            record_type_column = _detect_record_type_column(header.columns)
            if record_type_column:
                frame = pd.read_parquet(path, columns=[record_type_column])
                series = frame[record_type_column].fillna("").map(_display_record_type)
                counts = {label: int(quantity) for label, quantity in series.value_counts().items()}
    except Exception:
        counts = {}

    if not counts:
        file_key = _normalize_identifier(path.stem)
        try:
            row_count = int(round(float(fallback_row_count)))
        except Exception:
            row_count = None

        if row_count is not None:
            if "cadoc_1010_tipo4" in file_key:
                counts = {"Tipo 4": row_count}
            elif "cadoc_1010_operacoes" in file_key:
                counts = {"Tipo 1": row_count}
            elif "adip201" in file_key:
                counts = {"ADIP201": row_count}

    return counts


def enrich_summary_with_record_type_metrics(summary_df, input_dir):
    """Inclui quantidade de tipos e distribuição dos registros por tipo."""
    if summary_df is None or summary_df.empty:
        return pd.DataFrame() if summary_df is None else summary_df

    output = summary_df.copy()
    type_quantities = []
    type_distributions = []

    for _, row in output.iterrows():
        source_path = _match_input_file(input_dir, row)
        counts = (
            _read_record_type_counts(source_path, row.get("row_count"))
            if source_path is not None
            else {}
        )

        if counts:
            ordered = sorted(counts.items(), key=lambda item: _record_type_sort_key(item[0]))
            type_quantities.append(len(ordered))
            type_distributions.append(
                " | ".join(f"{label}: {quantity}" for label, quantity in ordered)
            )
        else:
            type_quantities.append("N/A")
            type_distributions.append("N/A")

    output["record_type_count"] = type_quantities
    output["records_by_type"] = type_distributions
    return output


def _recalculate_profile_classification(work, profile):
    if work.empty:
        return work

    if profile == "bacen_catalog":
        if "status" not in work.columns:
            work["status"] = "Não avaliado"
        for idx, row in work.iterrows():
            evaluated = pd.to_numeric(
                pd.Series([row.get("evaluated_datasets")]), errors="coerce"
            ).iloc[0]
            score = pd.to_numeric(
                pd.Series([row.get("average_score")]), errors="coerce"
            ).iloc[0]
            if pd.isna(evaluated) or evaluated <= 0 or pd.isna(score):
                work.at[idx, "status"] = "Não avaliado"
            else:
                work.at[idx, "status"] = classify_score(score)

    elif profile == "bacen_detail":
        if "status" not in work.columns:
            work["status"] = "Não avaliado"
        for idx, row in work.iterrows():
            original_status = normalize_status_label(row.get("status"))
            raw_score = pd.to_numeric(
                pd.Series([row.get("raw_score")]), errors="coerce"
            ).iloc[0]
            applicable = row.get("applicable", True)

            if _is_false_like(applicable):
                work.at[idx, "status"] = "Não aplicável"
            elif original_status == "Não avaliado" or pd.isna(raw_score):
                work.at[idx, "status"] = "Não avaliado"
            else:
                work.at[idx, "status"] = classify_score(raw_score)

    elif profile in {"attention", "top_bottom"}:
        if "score" in work.columns:
            work["classification"] = pd.to_numeric(
                work["score"], errors="coerce"
            ).apply(classify_score)

    return work


def _select_profile_columns(df, profile=None):
    if df is None:
        return pd.DataFrame()

    work = df.copy()

    # O resumo executivo mostra uma única aderência e uma única classificação.
    if profile == "summary":
        score_column = next(
            (
                column
                for column in ["score_bacen", "score", "score_final", "score_base"]
                if column in work.columns
            ),
            None,
        )
        selected = [
            column
            for column in [
                "dataset_label", "row_count", "column_count",
                "record_type_count", "records_by_type",
            ]
            if column in work.columns
        ]
        if score_column:
            work["_summary_score"] = pd.to_numeric(
                work[score_column], errors="coerce"
            )
            work["_summary_classification"] = work["_summary_score"].apply(
                classify_score
            )
            selected.extend(["_summary_score", "_summary_classification"])
        return work[selected] if selected else work

    work = _recalculate_profile_classification(work, profile)
    preferred = TABLE_PROFILES.get(profile)
    if preferred is None:
        return work

    selected = [column for column in preferred if column in work.columns]
    return work[selected] if selected else work


def _format_number_pt_br(value, decimals=2):
    try:
        number = float(value)
        if pd.isna(number):
            return ""
    except Exception:
        return safe_str(value)
    text = f"{number:,.{decimals}f}"
    return text.replace(",", "X").replace(".", ",").replace("X", ".")


def _format_integer_pt_br(value):
    try:
        number = float(value)
        if pd.isna(number):
            return ""
        return f"{int(round(number)):,}".replace(",", ".")
    except Exception:
        return safe_str(value)


def _percentage_to_fraction(value):
    try:
        number = float(value)
        if pd.isna(number):
            return None
    except Exception:
        return value
    return number / 100.0 if abs(number) > 1 else number


def _format_percentage_pt_br(value):
    fraction = _percentage_to_fraction(value)
    if fraction is None:
        return ""
    try:
        return f"{_format_number_pt_br(float(fraction) * 100, 2)}%"
    except Exception:
        return safe_str(value)


def _format_date_pt_br(value):
    if value is None or safe_str(value) == "":
        return ""
    try:
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.isna(parsed):
            return safe_str(value)
        return parsed.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return safe_str(value)


def _excel_series_with_na(numeric_series, divisor=1.0):
    """
    Cria uma série mista para o Excel, preservando números e exibindo N/A.

    O pandas usa o dtype anulável Float64 em algumas leituras do DuckDB.
    Esse dtype não aceita texto diretamente. A conversão explícita para
    object evita o erro: Invalid value 'N/A' for dtype 'Float64'.
    """
    numeric = pd.to_numeric(numeric_series, errors="coerce")
    result = (numeric / divisor).astype("object")
    result.loc[numeric.isna()] = "N/A"
    return result


def _excel_percentage_series_with_na(series):
    values = []
    for value in series:
        if _is_missing_metric_value(value):
            values.append("N/A")
        else:
            converted = _percentage_to_fraction(value)
            values.append("N/A" if converted is None else converted)
    return pd.Series(values, index=series.index, dtype="object")


def prepare_dataframe_for_excel(df, profile=None):
    """
    Traduz e formata a camada de apresentação sem alterar os dados técnicos.
    Notas de 0 a 10 viram frações de 0 a 1 nas abas executivas, permitindo
    que o Excel exiba corretamente percentuais de aderência.

    Métricas não mensuradas são apresentadas como N/A. Um zero realmente
    mensurado continua sendo exportado como 0,00%.
    """
    work = _select_profile_columns(df, profile)
    if work.empty:
        return work

    original_columns = list(work.columns)

    for column in original_columns:
        if _is_adherence_score_column(profile, column):
            work[column] = _excel_series_with_na(work[column], divisor=10.0)
        elif _is_weighted_contribution_column(profile, column):
            # weighted_score é contribuição para o total, não uma classificação isolada.
            work[column] = _excel_series_with_na(work[column], divisor=10.0)
        elif column in PERCENTAGE_COLUMNS:
            work[column] = _excel_percentage_series_with_na(work[column])

    work.columns = _unique_labels([
        _profile_column_label(profile, column)
        for column in original_columns
    ])

    for column in work.columns:
        if pd.api.types.is_object_dtype(work[column]) or pd.api.types.is_bool_dtype(work[column]):
            work[column] = work[column].map(_translate_cell_value)

    return work


def _cell_class(profile, original_column, raw_value):
    classes = []

    if _is_adherence_score_column(profile, original_column) and _is_missing_metric_value(raw_value):
        classes.append("cell-score-na")
        classification = None
    elif _is_adherence_score_column(profile, original_column):
        classification = classify_adherence(score_to_adherence(raw_value))
    elif original_column in SCORE_COLUMNS and not _is_weighted_contribution_column(
        profile, original_column
    ):
        classification = classify_score(raw_value)
    else:
        classification = None

    if classification == "Excelente":
        classes.append("cell-score-excellent")
    elif classification == "Bom":
        classes.append("cell-score-good")
    elif classification == "Satisfatório":
        classes.append("cell-score-adequate")
    elif classification == "Atenção":
        classes.append("cell-score-warning")
    elif classification == "Crítico":
        classes.append("cell-score-critical")

    if _is_weighted_contribution_column(profile, original_column):
        classes.append("cell-contribution")

    if original_column in LONG_TEXT_COLUMNS:
        classes.append("cell-long-text")

    layout_class = _column_layout_class(original_column)
    if layout_class:
        classes.append(layout_class)

    return " ".join(classes)


def _status_badge(value):
    translated = normalize_status_label(_translate_cell_value(value))
    upper = translated.upper()

    if upper == "EXCELENTE":
        css = "status-excellent"
    elif upper in {"BOM", "CONFORME", "APROVADO", "VÁLIDO", "VALIDO", "SUCESSO"}:
        css = "status-good"
    elif upper == "SATISFATÓRIO":
        css = "status-adequate"
    elif upper == "ATENÇÃO":
        css = "status-warning"
    elif upper in {"CRÍTICO", "REPROVADO", "INVÁLIDO", "INVALIDO", "ERRO"}:
        css = "status-critical"
    elif upper in {"NÃO AVALIADO", "NÃO APLICÁVEL"}:
        css = "status-neutral"
    else:
        return None

    return f'<span class="status-badge {css}">{html_escape(translated)}</span>'


def _format_html_value(profile, original_column, value):
    translated = _translate_cell_value(value)

    if (
        (_is_adherence_score_column(profile, original_column)
         or _is_weighted_contribution_column(profile, original_column))
        and _is_missing_metric_value(value)
    ):
        return "N/A", "Não foi possível mensurar esta métrica."

    if _is_adherence_score_column(profile, original_column):
        display = f"{_format_number_pt_br(score_to_adherence(value), 2)}%"
    elif _is_weighted_contribution_column(profile, original_column):
        display = f"{_format_number_pt_br(score_to_adherence(value), 2)} p.p."
    elif original_column in PERCENTAGE_COLUMNS:
        display = _format_percentage_pt_br(value)
    elif original_column in SCORE_COLUMNS:
        display = _format_number_pt_br(value, 2)
    elif original_column in COUNT_COLUMNS:
        display = _format_integer_pt_br(value)
    elif original_column in DATE_COLUMNS:
        display = _format_date_pt_br(value)
    elif original_column in {"status", "classification", "classification_bacen", "_summary_classification"}:
        display = normalize_status_label(value)
    elif isinstance(value, float):
        display = _format_number_pt_br(value, 4)
    else:
        display = safe_str(translated)

    full = display
    if original_column in LONG_TEXT_COLUMNS and len(display) > 140:
        display = display[:137].rstrip() + "..."

    return html_escape(display), html_escape(full)


def df_to_html(df, profile=None, max_rows=250):
    if df is None or df.empty:
        return '<div class="table-empty">Sem dados disponíveis para esta seção.</div>'

    work = _select_profile_columns(df, profile)
    if work.empty:
        return '<div class="table-empty">Sem colunas disponíveis para esta seção.</div>'

    total_rows = len(work)
    if total_rows > max_rows:
        work = work.head(max_rows)
        note = f"Exibindo os primeiros {max_rows} de {total_rows} registros."
    else:
        note = f"{total_rows} registro(s) exibido(s)."

    original_columns = list(work.columns)
    headers = _unique_labels([
        _profile_column_label(profile, column)
        for column in original_columns
    ])
    header_html = "".join(
        f'<th scope="col" class="{_column_layout_class(column)}">'
        f'{html_escape(label)}</th>'
        for column, label in zip(original_columns, headers)
    )

    rows_html = []
    for _, row in work.iterrows():
        cells = []
        for original_column in original_columns:
            raw = row.get(original_column)
            display, tooltip = _format_html_value(profile, original_column, raw)
            badge = (
                _status_badge(raw)
                if original_column in {
                    "status", "classification", "classification_bacen",
                    "_summary_classification",
                }
                else None
            )
            content = badge if badge else display
            cells.append(
                f'<td class="{_cell_class(profile, original_column, raw)}" '
                f'title="{tooltip}">{content}</td>'
            )
        rows_html.append("<tr>" + "".join(cells) + "</tr>")

    return f"""
    <div class="table-toolbar">
        <span class="table-count">{html_escape(note)}</span>
        <span class="table-hint">Role horizontalmente para visualizar todas as colunas.</span>
    </div>
    <div class="report-table-wrapper">
        <table class="report-table">
            <thead><tr>{header_html}</tr></thead>
            <tbody>{''.join(rows_html)}</tbody>
        </table>
    </div>
    """


def _style_excel_workbook(writer):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="0B1F3A")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="C5D0DE")
    excellent_fill = PatternFill("solid", fgColor="E7F5F0")
    good_fill = PatternFill("solid", fgColor="E8F7EF")
    adequate_fill = PatternFill("solid", fgColor="EAF2FF")
    warning_fill = PatternFill("solid", fgColor="FFF4E8")
    critical_fill = PatternFill("solid", fgColor="FDECEC")
    neutral_fill = PatternFill("solid", fgColor="F1F5F9")

    for sheet_index, ws in enumerate(writer.book.worksheets, start=1):
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(bottom=thin, left=thin, right=thin)
        ws.row_dimensions[1].height = 38

        if ws.max_row >= 2:
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            clean = ''.join(ch for ch in ws.title if ch.isalnum())[:18]
            table = Table(displayName=f"Tabela{sheet_index}_{clean}", ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)

        for col_idx in range(1, ws.max_column + 1):
            header = safe_str(ws.cell(1, col_idx).value)
            sample = [
                safe_str(ws.cell(row_idx, col_idx).value)
                for row_idx in range(1, min(ws.max_row, 120) + 1)
            ]
            longest = max([len(header), *(len(value) for value in sample)], default=10)
            width = (
                55
                if header in {
                    "Detalhes", "Evidências", "Evidência", "Descrição",
                    "Recomendação", "Colunas relacionadas", "Interpretação",
                }
                else min(max(longest + 2, 12), 36)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = width

            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row_idx, col_idx)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=thin, left=thin, right=thin)

                percentage_headers = {
                    "Aderência de qualidade", "Índice BACEN", "Média",
                    "Menor média avaliada", "Maior média avaliada",
                    "Aderência mínima esperada", "Aderência obtida",
                    "Contribuição ponderada", "Peso", "Melhor aderência",
                    "Pior aderência",
                }
                if header in percentage_headers or "Taxa" in header:
                    cell.number_format = "0.00%"
                elif "Pontuação" in header or "Nota técnica" in header:
                    cell.number_format = "0.00"
                elif any(term in header for term in [
                    "Total", "Quantidade", "Registros", "Conjuntos avaliados",
                    "Fontes avaliadas",
                ]):
                    cell.number_format = "#,##0"

                value = normalize_status_label(cell.value).upper()
                if value == "EXCELENTE":
                    cell.fill = excellent_fill
                elif value in {"BOM", "CONFORME", "APROVADO", "VÁLIDO", "VALIDO", "SUCESSO"}:
                    cell.fill = good_fill
                elif value == "SATISFATÓRIO":
                    cell.fill = adequate_fill
                elif value == "ATENÇÃO":
                    cell.fill = warning_fill
                elif value in {"CRÍTICO", "REPROVADO", "INVÁLIDO", "INVALIDO", "ERRO"}:
                    cell.fill = critical_fill
                elif value in {"NÃO AVALIADO", "NÃO APLICÁVEL"}:
                    cell.fill = neutral_fill


# =========================================================
# Utilidades
# =========================================================
def ensure_dir(path):
    p = Path(path)
    p.mkdir(parents=True, exist_ok=True)
    return p


def safe_str(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def html_escape(value):
    return (
        safe_str(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def table_exists(con, schema, table_name):
    row = con.execute(
        """
        SELECT 1
        FROM information_schema.tables
        WHERE table_schema = ? AND table_name = ?
        LIMIT 1
        """,
        [schema, table_name],
    ).fetchone()
    return row is not None


def fetchdf_safe(con, sql, params=None):
    try:
        if params is None:
            return con.execute(sql).fetchdf()
        return con.execute(sql, params).fetchdf()
    except Exception:
        return pd.DataFrame()


def detect_score_col(df):
    for col in ["score_bacen", "score", "score_final", "score_base"]:
        if col in df.columns:
            return col
    return None


def detect_classification_col(df):
    for col in ["classification_bacen", "classification", "classificacao"]:
        if col in df.columns:
            return col
    return None


def best_dataset_label(row):
    candidates = [
        row.get("dataset_label"),
        row.get("dataset_name"),
        row.get("dataset"),
        row.get("object_name"),
        row.get("source_name"),
        row.get("table_name"),
        row.get("file_name"),
        row.get("source_ref"),
        row.get("source"),
    ]
    for c in candidates:
        txt = safe_str(c)
        if txt:
            return txt
    return "Dataset não identificado"


def score_to_adherence(score):
    """Converte a nota técnica de 0 a 10 em aderência de 0% a 100%."""
    from decimal import Decimal, ROUND_HALF_UP

    try:
        value = float(score)
        if pd.isna(value):
            return 0.0
    except Exception:
        return 0.0

    value = max(0.0, min(value, 10.0))
    percentage = Decimal(str(value * 10.0)).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    return float(percentage)


def classify_adherence(percentage):
    """Classifica a aderência sem lacunas ou sobreposição entre as faixas."""
    try:
        value = float(percentage)
        if pd.isna(value):
            return "Sem classificação"
    except Exception:
        return "Sem classificação"

    value = max(0.0, min(value, 100.0))
    if value < 50.0:
        return "Crítico"
    if value < 70.0:
        return "Atenção"
    if value < 85.0:
        return "Satisfatório"
    if value < 95.0:
        return "Bom"
    return "Excelente"


def classify_score(score):
    return classify_adherence(score_to_adherence(score))


def status_color(status):
    mapping = {
        "Excelente": "#0F7B5B",
        "Bom": "#2EAD67",
        "Satisfatório": "#0F5DFF",
        "Atenção": "#F28C28",
        "Crítico": "#D64545",
        "Sem classificação": "#6B7280",
    }
    return mapping.get(status, "#6B7280")


def chart_color_for_adherence(value):
    """Usa a mesma paleta em cartões, gráficos, tabelas e legendas."""
    return status_color(classify_adherence(value))


def get_latest_run_id(con, schema):
    for table in ["dq_table_scores_u", "dq_table_scores_u_rules", "dq_column_scores_u"]:
        if not table_exists(con, schema, table):
            continue
        row = con.execute(
            f"""
            SELECT run_id
            FROM {schema}.{table}
            WHERE run_id IS NOT NULL
            ORDER BY run_id DESC
            LIMIT 1
            """
        ).fetchone()
        if row and row[0]:
            return str(row[0])
    return None


# =========================================================
# Leitura de dados
# =========================================================
def load_summary(con, schema, run_id):
    for table in ["dq_table_scores_u", "dq_table_scores_u_rules"]:
        if not table_exists(con, schema, table):
            continue
        df = fetchdf_safe(con, f"SELECT * FROM {schema}.{table} WHERE run_id = ?", [run_id])
        if not df.empty:
            out = df.copy()
            if "dataset_label" not in out.columns:
                out["dataset_label"] = out.apply(best_dataset_label, axis=1)
            score_col = detect_score_col(out)
            if score_col:
                out[score_col] = pd.to_numeric(out[score_col], errors="coerce").fillna(0)
            class_col = detect_classification_col(out)
            if not class_col and score_col:
                out["classification"] = out[score_col].apply(classify_score)
            return out
    return pd.DataFrame()


def load_detail(con, schema, run_id):
    if not table_exists(con, schema, "dq_column_scores_u"):
        return pd.DataFrame()
    df = fetchdf_safe(con, f"SELECT * FROM {schema}.dq_column_scores_u WHERE run_id = ?", [run_id])
    if df.empty:
        return df
    out = df.copy()
    if "dataset_label" not in out.columns:
        out["dataset_label"] = out.apply(best_dataset_label, axis=1)
    score_col = detect_score_col(out)
    if score_col:
        out[score_col] = pd.to_numeric(out[score_col], errors="coerce").fillna(0)
    class_col = detect_classification_col(out)
    if not class_col and score_col:
        out["classification"] = out[score_col].apply(classify_score)
    return out


def derive_summary_from_detail(detail_df, run_id):
    if detail_df.empty:
        return pd.DataFrame()
    score_col = detect_score_col(detail_df)
    if not score_col:
        return pd.DataFrame()

    grouped = (
        detail_df.groupby("dataset_label", dropna=False)[score_col]
        .mean()
        .reset_index()
        .rename(columns={score_col: "score"})
    )
    grouped["run_id"] = run_id
    grouped["classification"] = grouped["score"].apply(classify_score)
    return grouped


def load_bacen_dimensions(con, schema, run_id):
    if not table_exists(con, schema, "dq_bacen_summary"):
        return pd.DataFrame()
    df = fetchdf_safe(con, f"SELECT * FROM {schema}.dq_bacen_summary WHERE run_id = ?", [run_id])
    if df.empty:
        return df
    out = df.copy()
    if "dataset_label" not in out.columns:
        out["dataset_label"] = out.apply(best_dataset_label, axis=1)
    return out


def load_bacen_dimension_detail(con, schema, run_id):
    """Carrega uma linha por dataset e dimensão, incluindo peso, tipo e evidência."""
    if not table_exists(con, schema, "dq_bacen_dimension_scores"):
        return pd.DataFrame()
    df = fetchdf_safe(con, f"SELECT * FROM {schema}.dq_bacen_dimension_scores WHERE run_id = ?", [run_id])
    if df.empty:
        return df
    out = df.copy()
    if "dataset_label" not in out.columns:
        out["dataset_label"] = out.apply(best_dataset_label, axis=1)
    for col in ["weight", "minimum_score", "raw_score", "weighted_score"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")
    return out


BACEN_DIMENSIONS = [
    ("acessibilidade", "Acessibilidade", "Disponibilidade e acesso autorizado aos dados, considerando perfis, segregação de funções e condições de consulta."),
    ("acuracia", "Acurácia", "Correspondência dos dados com a realidade representada, combinando validade, unicidade, regras de negócio e reconciliação."),
    ("adaptabilidade", "Adaptabilidade", "Capacidade de adequar dados, estruturas e controles a mudanças regulatórias, tecnológicas e de negócio."),
    ("clareza", "Clareza", "Facilidade de compreensão dos dados por meio de nomes, definições, dicionário, metadados e documentação."),
    ("comparabilidade", "Comparabilidade", "Possibilidade de comparar informações entre períodos, fontes, produtos ou entidades usando critérios padronizados."),
    ("completude", "Completude", "Nível de preenchimento dos campos obrigatórios e relevantes, com controle de valores nulos, vazios ou ausentes."),
    ("confiabilidade", "Confiabilidade", "Grau de confiança no uso do dado, considerando consistência, integridade e cumprimento das regras aplicáveis."),
    ("consistencia", "Consistência", "Coerência dos valores, tipos, formatos e regras entre registros, colunas e fontes relacionadas."),
    ("integridade", "Integridade", "Preservação das relações, chaves e vínculos esperados, evitando registros órfãos ou relações quebradas."),
    ("rastreabilidade", "Rastreabilidade", "Capacidade de identificar origem, execução, fonte, regra aplicada, transformação e evidência associada ao dado."),
    ("relevancia", "Relevância", "Adequação do conteúdo ao objetivo regulatório, operacional ou analítico para o qual os dados serão utilizados."),
    ("tempestividade", "Tempestividade", "Disponibilidade e atualização dos dados dentro do prazo necessário para seu uso e para o cumprimento regulatório."),
]


def bacen_dimension_description_rows():
    rows = []
    for _code, label, description in BACEN_DIMENSIONS:
        rows.append(
            '<tr><td style="padding:16px 18px; border-top:1px solid #E5E7EB; font-weight:700; color:#0F172A; background:#F8FAFC;">'
            + html_escape(label)
            + '</td><td style="padding:16px 18px; border-top:1px solid #E5E7EB; color:#374151;">'
            + html_escape(description)
            + '</td></tr>'
        )
    return "".join(rows)


def build_bacen_catalog(dim_df, detail_df):
    """Resume as 12 dimensões, cobertura, peso, tipo e evidência."""
    records = []
    detail = detail_df.copy() if detail_df is not None else pd.DataFrame()
    for code, label, description in BACEN_DIMENSIONS:
        col = f"dim_{code}"
        series = pd.Series(dtype=float)
        if dim_df is not None and not dim_df.empty and col in dim_df.columns:
            series = pd.to_numeric(dim_df[col], errors="coerce").dropna()
        part = pd.DataFrame()
        if not detail.empty and "dimension_code" in detail.columns:
            part = detail[detail["dimension_code"].astype(str).str.lower() == code]
        def unique_join(column):
            if part.empty or column not in part.columns:
                return ""
            vals = [safe_str(v) for v in part[column].dropna().tolist() if safe_str(v)]
            return " | ".join(dict.fromkeys(vals))
        weight = None
        if not part.empty and "weight" in part.columns:
            weights = pd.to_numeric(part["weight"], errors="coerce").dropna()
            if not weights.empty:
                weight = round(float(weights.iloc[0]), 4)
        average_score = round(float(series.mean()), 4) if not series.empty else None
        evaluated_datasets = int(series.count())
        records.append({
            "dimension_code": code,
            "dimension_name": label,
            "description": description,
            "average_score": average_score,
            "minimum_score_observed": round(float(series.min()), 4) if not series.empty else None,
            "maximum_score_observed": round(float(series.max()), 4) if not series.empty else None,
            "evaluated_datasets": evaluated_datasets,
            "total_datasets": int(len(dim_df)) if dim_df is not None else 0,
            "weight": weight,
            "evaluation_type": translate_evaluation_type(unique_join("evaluation_type")),
            "status": classify_score(average_score) if evaluated_datasets > 0 and average_score is not None else "Não avaliado",
            "evidence": unique_join("evidence"),
        })
    return pd.DataFrame(records)


def build_dimensions(summary_df, run_id):
    cols = [
        "run_id", "dataset_label", "dim_completude", "dim_unicidade",
        "dim_consistencia", "dim_validade", "dim_integridade_ref",
        "dim_freshness", "score_final"
    ]
    if summary_df.empty:
        return pd.DataFrame(columns=cols)

    out = pd.DataFrame()
    out["run_id"] = summary_df["run_id"] if "run_id" in summary_df.columns else run_id
    out["dataset_label"] = summary_df["dataset_label"]

    dim_map = {
        "completude": "dim_completude",
        "unicidade": "dim_unicidade",
        "consistencia": "dim_consistencia",
        "validade": "dim_validade",
        "integridade": "dim_integridade_ref",
        "freshness": "dim_freshness",
    }

    for src, dst in dim_map.items():
        if src in summary_df.columns:
            out[dst] = pd.to_numeric(summary_df[src], errors="coerce").fillna(0)
        else:
            out[dst] = float("nan")

    score_col = detect_score_col(summary_df)
    if score_col:
        out["score_final"] = pd.to_numeric(summary_df[score_col], errors="coerce").fillna(0)
    else:
        out["score_final"] = 0.0

    return out[cols]


def load_history(con, schema):
    for table in ["dq_table_scores_u", "dq_table_scores_u_rules"]:
        if not table_exists(con, schema, table):
            continue
        df = fetchdf_safe(con, f"SELECT * FROM {schema}.{table}")
        if df.empty:
            continue
        score_col = detect_score_col(df)
        if "run_id" not in df.columns or not score_col:
            continue
        tmp = df.copy()
        tmp[score_col] = pd.to_numeric(tmp[score_col], errors="coerce").fillna(0)
        history = (
            tmp.groupby("run_id", dropna=False)[score_col]
            .mean()
            .reset_index()
            .rename(columns={score_col: "avg_score"})
            .sort_values("run_id")
        )
        return history
    return pd.DataFrame(columns=["run_id", "avg_score"])


# =========================================================
# Preparação analítica
# =========================================================
def build_kpis(summary_df):
    if summary_df.empty:
        return {
            "datasets": 0,
            "score_medio": 0.0,
            "aderencia_percentual": 0.0,
            "classificacao": "Sem classificação",
            "melhor_dataset": "Sem dados",
            "melhor_score": 0.0,
            "melhor_aderencia": 0.0,
            "pior_dataset": "Sem dados",
            "pior_score": 0.0,
            "pior_aderencia": 0.0,
            "criticos": 0,
            "atencao": 0,
            "satisfatorios": 0,
            "bons": 0,
            "excelentes": 0,
            "aceitaveis": 0,
        }

    work = summary_df.copy()
    score_col = detect_score_col(work)
    if not score_col:
        work["score"] = 0.0
        score_col = "score"

    work[score_col] = pd.to_numeric(work[score_col], errors="coerce").fillna(0)
    work["classification_adherence"] = work[score_col].apply(classify_score)

    avg_score = round(float(work[score_col].mean()), 4)
    avg_adherence = score_to_adherence(avg_score)
    best = work.sort_values(score_col, ascending=False).iloc[0]
    worst = work.sort_values(score_col, ascending=True).iloc[0]

    classifications = work["classification_adherence"]

    return {
        "datasets": int(len(work)),
        "score_medio": round(avg_score, 2),
        "aderencia_percentual": avg_adherence,
        "classificacao": classify_adherence(avg_adherence),
        "melhor_dataset": safe_str(best.get("dataset_label")),
        "melhor_score": round(float(best.get(score_col, 0)), 2),
        "melhor_aderencia": score_to_adherence(best.get(score_col, 0)),
        "pior_dataset": safe_str(worst.get("dataset_label")),
        "pior_score": round(float(worst.get(score_col, 0)), 2),
        "pior_aderencia": score_to_adherence(worst.get(score_col, 0)),
        "criticos": int((classifications == "Crítico").sum()),
        "atencao": int((classifications == "Atenção").sum()),
        "satisfatorios": int((classifications == "Satisfatório").sum()),
        "bons": int((classifications == "Bom").sum()),
        "excelentes": int((classifications == "Excelente").sum()),
        "aceitaveis": int(
            classifications.isin(["Satisfatório", "Bom", "Excelente"]).sum()
        ),
    }


def build_top_bottom(summary_df):
    if summary_df.empty:
        return pd.DataFrame(), pd.DataFrame()

    work = summary_df.copy()
    score_col = detect_score_col(work)
    if not score_col:
        return pd.DataFrame(), pd.DataFrame()

    work[score_col] = pd.to_numeric(work[score_col], errors="coerce").fillna(0)
    work["classification"] = work[score_col].apply(classify_score)
    cols = ["dataset_label", score_col, "classification"]

    top = work.sort_values(score_col, ascending=False)[cols].head(5).copy()
    bottom = work.sort_values(score_col, ascending=True)[cols].head(5).copy()

    top = top.rename(columns={score_col: "score"})
    bottom = bottom.rename(columns={score_col: "score"})
    return top, bottom


def build_attention(detail_df):
    if detail_df.empty:
        return pd.DataFrame()

    work = detail_df.copy()
    score_col = detect_score_col(work)
    if not score_col:
        return pd.DataFrame()

    work[score_col] = pd.to_numeric(work[score_col], errors="coerce").fillna(0)
    work["classification"] = work[score_col].apply(classify_score)
    cols = [
        column
        for column in [
            "dataset_label", "column_name", "dtype", score_col,
            "classification", "null_rate", "distinct_ratio", "violations",
        ]
        if column in work.columns
    ]
    out = work.sort_values(score_col, ascending=True)[cols].head(20).copy()
    return out.rename(columns={score_col: "score"})


def summary_message(kpis):
    adherence = _format_number_pt_br(kpis["aderencia_percentual"], 2)
    return (
        f"Foram avaliadas {kpis['datasets']} fontes de dados, com aderência média "
        f"de qualidade de {adherence}% e classificação geral "
        f"'{normalize_status_label(kpis['classificacao'])}'. O percentual representa o nível de atendimento "
        f"às regras, controles e dimensões de qualidade avaliadas. Foram identificadas "
        f"{kpis['criticos']} fontes críticas e {kpis['atencao']} fontes em atenção."
    )


def executive_list_html(kpis):
    worst_adherence = _format_number_pt_br(kpis["pior_aderencia"], 2)
    return f"""
    <ul class="exec">
        <li><strong>Fontes com aderência satisfatória ou superior:</strong> {kpis['aceitaveis']} de {kpis['datasets']}.</li>
        <li><strong>Fontes críticas:</strong> {kpis['criticos']}.</li>
        <li><strong>Fontes em atenção:</strong> {kpis['atencao']}.</li>
        <li><strong>Fontes satisfatórias:</strong> {kpis['satisfatorios']}.</li>
        <li><strong>Principal foco de correção:</strong> {html_escape(kpis['pior_dataset'])} ({worst_adherence}% de aderência).</li>
    </ul>
    """


ADHERENCE_RANGES = [
    {
        "percentual": "0,00% a 49,99%",
        "nota": "0,00 a 4,99",
        "classificacao": "Crítico",
        "resumo": (
            "Os dados apresentam baixa aderência às regras de qualidade e "
            "conformidade. É necessária revisão imediata dos controles, regras e registros."
        ),
    },
    {
        "percentual": "50,00% a 69,99%",
        "nota": "5,00 a 6,99",
        "classificacao": "Atenção",
        "resumo": (
            "Os dados atendem parcialmente aos critérios avaliados. É necessário "
            "revisar regras, corrigir inconsistências e fortalecer os controles."
        ),
    },
    {
        "percentual": "70,00% a 84,99%",
        "nota": "7,00 a 8,49",
        "classificacao": "Satisfatório",
        "resumo": (
            "A maior parte dos critérios de qualidade foi atendida, mas ainda existem "
            "pontos específicos que precisam de melhoria."
        ),
    },
    {
        "percentual": "85,00% a 94,99%",
        "nota": "8,50 a 9,49",
        "classificacao": "Bom",
        "resumo": (
            "Os dados apresentam alta aderência aos controles de qualidade, com poucas "
            "inconsistências e oportunidades pontuais de melhoria."
        ),
    },
    {
        "percentual": "95,00% a 100,00%",
        "nota": "9,50 a 10,00",
        "classificacao": "Excelente",
        "resumo": (
            "Os dados apresentam elevado nível de qualidade e conformidade. Recomenda-se "
            "manter o monitoramento e a governança contínua."
        ),
    },
]


def adherence_range_dataframe():
    return pd.DataFrame([
        {
            "Percentual de aderência": item["percentual"],
            "Nota técnica equivalente": item["nota"],
            "Classificação": item["classificacao"],
            "Interpretação": item["resumo"],
        }
        for item in ADHERENCE_RANGES
    ])


def adherence_range_html():
    rows = []
    for item in ADHERENCE_RANGES:
        badge = _status_badge(item["classificacao"]) or html_escape(item["classificacao"])
        rows.append(
            "<tr>"
            f"<td><strong>{html_escape(item['percentual'])}</strong></td>"
            f"<td>{html_escape(item['nota'])}</td>"
            f"<td>{badge}</td>"
            f"<td class=\"cell-long-text\">{html_escape(item['resumo'])}</td>"
            "</tr>"
        )

    return f"""
    <div class="range-note">
        O percentual de aderência representa o nível de atendimento dos dados às regras,
        controles e dimensões de qualidade avaliadas. A conversão utiliza a nota técnica
        de 0 a 10 multiplicada por 10.
    </div>
    <div class="report-table-wrapper range-table-wrapper">
        <table class="report-table range-table">
            <thead>
                <tr>
                    <th>Percentual de aderência</th>
                    <th>Nota técnica equivalente</th>
                    <th>Classificação</th>
                    <th>Interpretação</th>
                </tr>
            </thead>
            <tbody>{''.join(rows)}</tbody>
        </table>
    </div>
    """



def bacen_evaluation_glossary_dataframe():
    return pd.DataFrame([
        {
            "Tipo de avaliação": "Automatizado",
            "Definição": "Resultado calculado diretamente por regras, métricas e validações executadas pelo scanner.",
        },
        {
            "Tipo de avaliação": "Híbrido",
            "Definição": "Combina evidências automáticas do scanner com controles, documentos ou validações de governança.",
        },
        {
            "Tipo de avaliação": "Governança",
            "Definição": "Depende de evidências de processo, políticas, documentação, responsáveis, acessos, linhagem ou controles organizacionais.",
        },
    ])


def interpretation_criteria_dataframe():
    """Reúne faixas de aderência e tipos de avaliação em uma única referência."""
    rows = []

    for item in ADHERENCE_RANGES:
        rows.append({
            "Categoria": "Classificação da aderência",
            "Item": item["classificacao"],
            "Faixa de aderência": item["percentual"],
            "Nota técnica equivalente": item["nota"],
            "Interpretação": item["resumo"],
        })

    for _, row in bacen_evaluation_glossary_dataframe().iterrows():
        rows.append({
            "Categoria": "Tipo de avaliação",
            "Item": row["Tipo de avaliação"],
            "Faixa de aderência": "",
            "Nota técnica equivalente": "",
            "Interpretação": row["Definição"],
        })

    return pd.DataFrame(rows)


def bacen_evaluation_glossary_html():
    return """
    <div class="glossary-block">
        <h3>Como interpretar o tipo de avaliação</h3>
        <div class="glossary-grid">
            <div class="glossary-card"><strong>Automatizado</strong><span>Calculado diretamente por regras, métricas e validações executadas pelo scanner.</span></div>
            <div class="glossary-card"><strong>Híbrido</strong><span>Combina evidências automáticas com controles, documentos ou validações de governança.</span></div>
            <div class="glossary-card"><strong>Governança</strong><span>Depende de políticas, documentação, responsáveis, acessos, linhagem e controles organizacionais.</span></div>
        </div>
        <h3 class="classification-title">Como interpretar a classificação</h3>
        <div class="classification-legend">
            <span class="status-badge status-neutral">Não avaliado</span><span>Sem evidência suficiente para atribuir aderência.</span>
            <span class="status-badge status-critical">Crítico</span><span>De 0,00% a 49,99%.</span>
            <span class="status-badge status-warning">Atenção</span><span>De 50,00% a 69,99%.</span>
            <span class="status-badge status-adequate">Satisfatório</span><span>De 70,00% a 84,99%.</span>
            <span class="status-badge status-good">Bom</span><span>De 85,00% a 94,99%.</span>
            <span class="status-badge status-excellent">Excelente</span><span>De 95,00% a 100,00%.</span>
        </div>
    </div>
    """


# =========================================================
# Análise GenAI das métricas mensuradas
# =========================================================
GENAI_DEFAULT_PROVIDER = "Ollama"
GENAI_DEFAULT_MODEL = "llama3.2:3b"


def _env_flag(name, default=False):
    raw = os.getenv(name)
    if raw is None:
        return default
    return raw.strip().lower() in {"1", "true", "sim", "yes", "on"}


def _safe_float_or_none(value):
    try:
        number = float(value)
        if pd.isna(number):
            return None
        return number
    except Exception:
        return None


def _is_applicable_for_genai(value):
    if value is None:
        return True
    return safe_str(value).strip().lower() not in {
        "false", "0", "não", "nao", "n", "no",
        "não aplicável", "nao aplicavel", "n/a",
    }


def _is_measured_status(value):
    normalized = normalize_status_label(value)
    return normalized not in {
        "Não avaliado", "Não aplicável", "Sem classificação", "N/A"
    }


def _dimension_label_from_code(code):
    normalized = _normalize_identifier(code)
    for dimension_code, label, _description in BACEN_DIMENSIONS:
        if normalized == _normalize_identifier(dimension_code):
            return label
    return safe_str(code).replace("_", " ").title()


def build_genai_metric_payload(run_id, kpis, summary_df, dim_df, bacen_detail_df, semantic_df):
    """Monta pacote agregado sem PII, apenas com métricas mensuradas."""
    sources = []
    score_col = detect_score_col(summary_df) if summary_df is not None else None
    if summary_df is not None and not summary_df.empty and score_col:
        for _, row in summary_df.iterrows():
            score = _safe_float_or_none(row.get(score_col))
            if score is None:
                continue
            sources.append({
                "conjunto_de_dados": best_dataset_label(row),
                "aderencia_percentual": score_to_adherence(score),
                "classificacao": classify_score(score),
                "quantidade_registros": _safe_float_or_none(row.get("row_count")),
                "quantidade_colunas": _safe_float_or_none(row.get("column_count")),
            })

    measured_dimensions = []
    excluded_dimensions = 0
    if bacen_detail_df is not None and not bacen_detail_df.empty:
        for _, row in bacen_detail_df.iterrows():
            raw_score = _safe_float_or_none(row.get("raw_score"))
            applicable = _is_applicable_for_genai(row.get("applicable", True))
            measured_status = _is_measured_status(row.get("status"))
            if raw_score is None or not applicable or not measured_status:
                excluded_dimensions += 1
                continue
            measured_dimensions.append({
                "conjunto_de_dados": safe_str(row.get("dataset_label")) or "Não identificado",
                "dimensao": safe_str(row.get("dimension_name")) or _dimension_label_from_code(row.get("dimension_code")),
                "aderencia_percentual": score_to_adherence(raw_score),
                "classificacao": classify_score(raw_score),
                "tipo_avaliacao": translate_evaluation_type(row.get("evaluation_type")),
                "evidencia": safe_str(row.get("evidence"))[:500],
            })
    elif dim_df is not None and not dim_df.empty:
        dimension_columns = [column for column in dim_df.columns if str(column).startswith("dim_")]
        for _, row in dim_df.iterrows():
            dataset = safe_str(row.get("dataset_label")) or "Não identificado"
            for column in dimension_columns:
                score = _safe_float_or_none(row.get(column))
                if score is None:
                    excluded_dimensions += 1
                    continue
                measured_dimensions.append({
                    "conjunto_de_dados": dataset,
                    "dimensao": COLUMN_LABELS_PT_BR.get(column, column.replace("dim_", "").replace("_", " ").title()),
                    "aderencia_percentual": score_to_adherence(score),
                    "classificacao": classify_score(score),
                    "tipo_avaliacao": "Automatizado",
                    "evidencia": "Métrica calculada pelo scanner de Data Quality.",
                })

    semantic_metrics = []
    excluded_semantic = 0
    if semantic_df is not None and not semantic_df.empty:
        for _, row in semantic_df.iterrows():
            valid_rate = _safe_float_or_none(row.get("valid_rate"))
            evaluated_count = _safe_float_or_none(row.get("evaluated_count"))
            status = normalize_status_label(row.get("status"))
            if valid_rate is None or evaluated_count is None or evaluated_count <= 0:
                excluded_semantic += 1
                continue
            if status in {"Não avaliado", "Não aplicável", "Sem classificação"}:
                excluded_semantic += 1
                continue
            percentage = valid_rate * 100.0 if abs(valid_rate) <= 1 else valid_rate
            semantic_metrics.append({
                "conjunto_de_dados": safe_str(row.get("object_name")) or "Não identificado",
                "validacao": safe_str(row.get("validation_name")) or safe_str(row.get("validation_code")),
                "coluna": safe_str(row.get("column_name")),
                "aderencia_percentual": round(percentage, 2),
                "registros_avaliados": int(evaluated_count),
                "registros_invalidos": int(_safe_float_or_none(row.get("invalid_count")) or 0),
                "fonte_referencia": safe_str(row.get("reference_source")),
            })

    return {
        "run_id": safe_str(run_id),
        "aderencia_geral_percentual": round(float(kpis.get("aderencia_percentual", 0)), 2),
        "classificacao_geral": normalize_status_label(kpis.get("classificacao")),
        "fontes_avaliadas": int(kpis.get("datasets", 0)),
        "fontes": sources,
        "dimensoes_mensuradas": measured_dimensions,
        "validacoes_semanticas_mensuradas": semantic_metrics,
        "quantidade_metricas_consideradas": len(measured_dimensions) + len(semantic_metrics),
        "quantidade_metricas_nao_mensuradas_excluidas": excluded_dimensions + excluded_semantic,
        "regra_de_escopo": (
            "A análise considera somente métricas efetivamente mensuradas. "
            "Itens N/A, não avaliados, não aplicáveis ou sem evidência foram excluídos. "
            "Resultados mensurados iguais a zero permanecem no escopo."
        ),
    }


def _genai_schema():
    """Schema enxuto para respostas mais rápidas e fáceis de compreender."""
    return {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "resumo_executivo": {"type": "string"},
            "pontos_positivos": {
                "type": "array",
                "items": {"type": "string"},
                "maxItems": 3,
            },
            "riscos_prioritarios": {
                "type": "array",
                "maxItems": 4,
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "prioridade": {
                            "type": "string",
                            "enum": ["Alta", "Média", "Baixa"],
                        },
                        "titulo": {"type": "string"},
                        "descricao": {"type": "string"},
                    },
                    "required": ["prioridade", "titulo", "descricao"],
                },
            },
            "recomendacoes": {
                "type": "array",
                "items": {"type": "string"},
                "maxItems": 4,
            },
        },
        "required": [
            "resumo_executivo",
            "pontos_positivos",
            "riscos_prioritarios",
            "recomendacoes",
        ],
    }


def _metric_display_name(item):
    return safe_str(item.get("nome")) or "Métrica não identificada"


def _deterministic_genai_fallback(payload, reason=""):
    """
    Cria um resumo simples e útil quando a LLM local não responde.

    A falha técnica é preservada somente nos metadados/JSON e não é exibida
    no relatório executivo.
    """
    all_metrics = []

    for item in payload.get("dimensoes_mensuradas", []):
        all_metrics.append({
            "nome": f"{item.get('dimensao')} — {item.get('conjunto_de_dados')}",
            "aderencia": item.get("aderencia_percentual"),
            "classificacao": item.get("classificacao"),
        })

    for item in payload.get("validacoes_semanticas_mensuradas", []):
        adherence = item.get("aderencia_percentual")
        all_metrics.append({
            "nome": f"{item.get('validacao')} — {item.get('conjunto_de_dados')}",
            "aderencia": adherence,
            "classificacao": classify_adherence(adherence),
        })

    measured = [
        item for item in all_metrics
        if _safe_float_or_none(item.get("aderencia")) is not None
    ]
    measured.sort(key=lambda item: float(item["aderencia"]))

    adherence = float(payload.get("aderencia_geral_percentual", 0) or 0)
    classification = safe_str(
        payload.get("classificacao_geral")
    ) or "Sem classificação"
    count = int(payload.get("quantidade_metricas_consideradas", 0) or 0)

    interpretation = {
        "Crítico": (
            "O resultado indica problemas importantes e exige correções prioritárias "
            "antes que os dados sejam usados em decisões relevantes."
        ),
        "Atenção": (
            "Parte das regras foi atendida, mas ainda existem falhas relevantes que "
            "precisam ser corrigidas."
        ),
        "Satisfatório": (
            "A maior parte das regras foi atendida, porém alguns pontos ainda precisam "
            "de correção para aumentar a confiança nos dados."
        ),
        "Bom": (
            "Os dados apresentam boa qualidade, com poucos pontos de melhoria."
        ),
        "Excelente": (
            "Os dados apresentam alto nível de qualidade e podem ser mantidos sob "
            "monitoramento contínuo."
        ),
    }.get(classification, "O resultado deve ser analisado em conjunto com os riscos destacados.")

    summary = (
        f"De forma geral, os dados avaliados alcançaram "
        f"{_format_number_pt_br(adherence, 2)}% de aderência, resultado considerado "
        f"{classification}. {interpretation} A análise considerou {count} métricas "
        f"que puderam ser efetivamente medidas."
    )

    best_candidates = [
        item for item in sorted(
            measured,
            key=lambda row: float(row["aderencia"]),
            reverse=True,
        )
        if float(item["aderencia"]) >= 85
    ][:3]

    if not best_candidates:
        best_candidates = sorted(
            measured,
            key=lambda row: float(row["aderencia"]),
            reverse=True,
        )[:2]

    positives = []
    for item in best_candidates:
        positives.append(
            f"{_metric_display_name(item)} apresentou "
            f"{_format_number_pt_br(item['aderencia'], 2)}% de aderência "
            f"e foi classificada como {item['classificacao']}."
        )

    risk_candidates = [
        item for item in measured
        if float(item["aderencia"]) < 70
    ][:4]

    risks = []
    for item in risk_candidates:
        value = float(item["aderencia"])
        priority = "Alta" if value < 50 else "Média"
        plain_language = (
            "Esse resultado indica uma falha importante e deve ser corrigido primeiro."
            if value < 50
            else "Esse resultado merece atenção e deve ser revisado na próxima etapa de correção."
        )
        risks.append({
            "prioridade": priority,
            "titulo": _metric_display_name(item),
            "descricao": (
                f"A aderência foi de {_format_number_pt_br(value, 2)}%. "
                f"{plain_language}"
            ),
        })

    recommendations = []
    for risk in risks[:3]:
        recommendations.append(
            f"Revisar {risk['titulo']} e corrigir a regra, o dado de origem ou o processo "
            "responsável pelo resultado."
        )

    if not recommendations:
        recommendations.append(
            "Manter o monitoramento das métricas e repetir a avaliação após novas cargas de dados."
        )
    else:
        recommendations.append(
            "Executar novamente o scanner após as correções para confirmar a melhoria da aderência."
        )

    return {
        "resumo_executivo": summary,
        "pontos_positivos": positives or [
            "Não houve resultados suficientes para destacar um ponto positivo."
        ],
        "riscos_prioritarios": risks,
        "recomendacoes": recommendations[:4],
    }


def _validate_genai_output(result):
    required = {
        "resumo_executivo",
        "pontos_positivos",
        "riscos_prioritarios",
        "recomendacoes",
    }
    if not isinstance(result, dict) or not required.issubset(result):
        raise ValueError("A resposta GenAI não possui o formato esperado.")
    return result


def _extract_json_from_text(text_value):
    """Extrai um objeto JSON mesmo quando o modelo adiciona cercas Markdown."""
    raw = safe_str(text_value)
    if not raw:
        raise ValueError("A LLM retornou uma resposta vazia.")

    try:
        return json.loads(raw)
    except Exception:
        pass

    cleaned = re.sub(r"^```(?:json)?\\s*", "", raw.strip(), flags=re.IGNORECASE)
    cleaned = re.sub(r"\\s*```$", "", cleaned.strip())
    try:
        return json.loads(cleaned)
    except Exception:
        pass

    start_pos = cleaned.find("{")
    end_pos = cleaned.rfind("}")
    if start_pos >= 0 and end_pos > start_pos:
        return json.loads(cleaned[start_pos:end_pos + 1])

    raise ValueError("Não foi possível interpretar a resposta da LLM como JSON.")


def _genai_instructions():
    return (
        "Você analisa resultados de qualidade de dados para pessoas que não são técnicas. "
        "Use somente o JSON fornecido e considere apenas as métricas mensuradas. "
        "Não mencione métricas N/A, não avaliadas, não aplicáveis ou excluídas. "
        "Um resultado realmente medido como 0% deve ser tratado como crítico. "
        "Não altere números, não invente evidências e não declare conformidade regulatória. "
        "Escreva em português do Brasil, com frases curtas, linguagem simples e prática. "
        "O resumo deve ter no máximo três frases. Destaque até três pontos positivos. "
        "Liste como riscos somente métricas Críticas ou em Atenção. "
        "As recomendações devem explicar claramente o que precisa ser feito. "
        "Não descreva o funcionamento da LLM, não mencione fallback, timeout, prompt, API ou detalhes técnicos. "
        "Retorne somente um objeto JSON válido de acordo com o schema informado."
    )


def _call_ollama(model, payload):
    """Executa a análise na API local do Ollama, com saída curta e uma repetição segura."""
    base_url = os.getenv(
        "MJV_OLLAMA_BASE_URL", "http://localhost:11434"
    ).strip().rstrip("/")
    timeout_seconds = int(os.getenv("MJV_OLLAMA_TIMEOUT_SECONDS", "420"))
    max_attempts = int(os.getenv("MJV_OLLAMA_MAX_ATTEMPTS", "2"))

    compact_payload = json.dumps(
        payload,
        ensure_ascii=False,
        separators=(",", ":"),
    )
    request_body = {
        "model": model,
        "messages": [
            {"role": "system", "content": _genai_instructions()},
            {"role": "user", "content": compact_payload},
        ],
        "stream": False,
        "keep_alive": "10m",
        "format": _genai_schema(),
        "options": {
            "temperature": 0.1,
            "num_predict": 480,
            "num_ctx": 4096,
        },
    }

    encoded = json.dumps(request_body, ensure_ascii=False).encode("utf-8")
    last_error = None

    for attempt in range(1, max_attempts + 1):
        request = urllib_request.Request(
            f"{base_url}/api/chat",
            data=encoded,
            headers={"Content-Type": "application/json; charset=utf-8"},
            method="POST",
        )
        try:
            with urllib_request.urlopen(
                request,
                timeout=timeout_seconds,
            ) as response:
                response_data = json.loads(response.read().decode("utf-8"))

            content = (
                response_data.get("message", {}).get("content")
                or response_data.get("response")
                or ""
            )
            result = _validate_genai_output(_extract_json_from_text(content))
            request_id = safe_str(
                response_data.get("created_at") or response_data.get("model")
            )
            return result, request_id

        except urllib_error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(
                f"Ollama respondeu HTTP {exc.code}: {body[:300]}"
            ) from exc
        except urllib_error.URLError as exc:
            last_error = RuntimeError(
                "Não foi possível conectar ao Ollama em "
                f"{base_url}. Confirme se o Ollama está em execução."
            )
        except (TimeoutError, OSError) as exc:
            last_error = exc
        except Exception as exc:
            last_error = exc

        if attempt < max_attempts:
            time.sleep(2)

    if last_error is not None:
        raise last_error
    raise RuntimeError("A análise local não retornou uma resposta válida.")


def _call_openai(model, payload):
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY não configurada.")

    from openai import OpenAI

    client = OpenAI(api_key=api_key)
    response = client.responses.create(
        model=model,
        instructions=_genai_instructions(),
        input=json.dumps(payload, ensure_ascii=False),
        text={
            "format": {
                "type": "json_schema",
                "name": "mjv_data_quality_genai_analysis",
                "strict": True,
                "schema": _genai_schema(),
            }
        },
    )
    result = _validate_genai_output(json.loads(response.output_text))
    return result, safe_str(getattr(response, "_request_id", ""))


def generate_genai_analysis(run_id, kpis, summary_df, dim_df, bacen_detail_df, semantic_df):
    """Executa a IA local e mantém um resumo automático caso ela não responda."""
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except Exception:
        pass

    provider_raw = (
        os.getenv("MJV_GENAI_PROVIDER", GENAI_DEFAULT_PROVIDER).strip()
        or GENAI_DEFAULT_PROVIDER
    )
    provider_key = provider_raw.lower()
    model = (
        os.getenv("MJV_GENAI_MODEL", GENAI_DEFAULT_MODEL).strip()
        or GENAI_DEFAULT_MODEL
    )
    enabled = _env_flag("MJV_GENAI_ENABLED", default=False)
    payload = build_genai_metric_payload(
        run_id, kpis, summary_df, dim_df, bacen_detail_df, semantic_df
    )

    if provider_key in {"ollama", "local", "ollama local", "ollama-local"}:
        provider_display = "Ollama — execução local"
    elif provider_key == "openai":
        provider_display = "OpenAI"
    else:
        provider_display = provider_raw

    metadata = {
        "enabled": enabled,
        "provider": provider_display,
        "model": model,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "status": "não executada",
        "request_id": "",
        "metricas_consideradas": payload["quantidade_metricas_consideradas"],
        "metricas_excluidas": payload["quantidade_metricas_nao_mensuradas_excluidas"],
        "scope_note": payload["regra_de_escopo"],
        "diagnostico_tecnico": "",
    }

    if not enabled:
        analysis = _deterministic_genai_fallback(payload)
        metadata["status"] = "resumo automático"
        metadata["diagnostico_tecnico"] = (
            "GenAI desativada. Defina MJV_GENAI_ENABLED=true para ativar a LLM."
        )
        return {"metadata": metadata, "payload": payload, "analysis": analysis}

    if payload["quantidade_metricas_consideradas"] <= 0:
        analysis = _deterministic_genai_fallback(payload)
        metadata["status"] = "sem métricas mensuradas"
        metadata["diagnostico_tecnico"] = (
            "Nenhuma métrica mensurada estava disponível para análise."
        )
        return {"metadata": metadata, "payload": payload, "analysis": analysis}

    try:
        if provider_key in {"ollama", "local", "ollama local", "ollama-local"}:
            result, request_id = _call_ollama(model, payload)
        elif provider_key == "openai":
            result, request_id = _call_openai(model, payload)
        else:
            raise RuntimeError(
                f"Provedor GenAI não suportado: {provider_raw}. Use ollama ou openai."
            )

        metadata["status"] = "executada com sucesso"
        metadata["request_id"] = request_id
        return {"metadata": metadata, "payload": payload, "analysis": result}

    except Exception as exc:
        analysis = _deterministic_genai_fallback(payload)
        metadata["status"] = "resumo automático"
        metadata["diagnostico_tecnico"] = (
            f"Falha na chamada GenAI: {type(exc).__name__}: "
            f"{safe_str(exc)[:400]}"
        )
        return {"metadata": metadata, "payload": payload, "analysis": analysis}


def genai_analysis_dataframe(genai_result):
    """Versão simples da análise para a aba Excel."""
    metadata = (genai_result or {}).get("metadata", {})
    analysis = (genai_result or {}).get("analysis", {})

    rows = [
        {
            "Seção": "Identificação",
            "Item": "Motor de análise",
            "Conteúdo": f"{metadata.get('provider', 'N/A')} • {metadata.get('model', 'N/A')}",
        },
        {
            "Seção": "Resumo",
            "Item": "Visão geral",
            "Conteúdo": analysis.get("resumo_executivo", ""),
        },
    ]

    for item in analysis.get("pontos_positivos", []):
        rows.append({
            "Seção": "Pontos positivos",
            "Item": "Destaque",
            "Conteúdo": item,
        })

    for item in analysis.get("recomendacoes", []):
        rows.append({
            "Seção": "Recomendações",
            "Item": "Ação sugerida",
            "Conteúdo": item,
        })

    for risk in analysis.get("riscos_prioritarios", []):
        rows.append({
            "Seção": "Riscos",
            "Item": f"{risk.get('prioridade', 'N/A')} — {risk.get('titulo', '')}",
            "Conteúdo": risk.get("descricao", ""),
        })

    return pd.DataFrame(rows)


def _genai_engine_icon_svg():
    """Ícone vetorial incorporado ao HTML; não depende de arquivo externo."""
    return """
    <svg viewBox="0 0 72 72" role="img" aria-label="Motor de inteligência artificial">
        <defs>
            <linearGradient id="aiEngineGradient" x1="0" y1="0" x2="1" y2="1">
                <stop offset="0%" stop-color="#0B1F3A"/>
                <stop offset="100%" stop-color="#0F5DFF"/>
            </linearGradient>
        </defs>
        <circle cx="36" cy="36" r="33" fill="url(#aiEngineGradient)"/>
        <path d="M36 17l4 2 5-1 3 4 5 2v5l3 4-3 4v5l-5 2-3 4-5-1-4 2-4-2-5 1-3-4-5-2v-5l-3-4 3-4v-5l5-2 3-4 5 1 4-2z" fill="none" stroke="#FFFFFF" stroke-width="2.4" stroke-linejoin="round"/>
        <circle cx="36" cy="33" r="8" fill="#FFFFFF" opacity="0.96"/>
        <circle cx="32" cy="31" r="1.8" fill="#0F5DFF"/>
        <circle cx="40" cy="31" r="1.8" fill="#0F5DFF"/>
        <circle cx="36" cy="37" r="1.8" fill="#0F5DFF"/>
        <path d="M32 31l4 6 4-6M36 41v5" fill="none" stroke="#FFFFFF" stroke-width="2" stroke-linecap="round"/>
    </svg>
    """


def genai_analysis_html(genai_result):
    """Bloco executivo simples: resumo, pontos positivos, recomendações e riscos."""
    metadata = (genai_result or {}).get("metadata", {})
    analysis = (genai_result or {}).get("analysis", {})
    status = safe_str(metadata.get("status"))
    success = status == "executada com sucesso"

    model_label = (
        f"LLM aplicada: {metadata.get('provider', 'N/A')} • {metadata.get('model', 'N/A')}"
        if success
        else f"LLM configurada: {metadata.get('provider', 'N/A')} • {metadata.get('model', 'N/A')}"
    )

    positives = "".join(
        f"<li>{html_escape(item)}</li>"
        for item in analysis.get("pontos_positivos", [])
    ) or "<li>Sem destaques disponíveis.</li>"

    recommendations = "".join(
        f"<li>{html_escape(item)}</li>"
        for item in analysis.get("recomendacoes", [])
    ) or "<li>Sem recomendações disponíveis.</li>"

    risks = []
    for risk in analysis.get("riscos_prioritarios", []):
        priority = safe_str(risk.get("prioridade")) or "N/A"
        priority_class = (
            "ai-risk-high" if priority == "Alta"
            else "ai-risk-medium" if priority == "Média"
            else "ai-risk-low"
        )
        risks.append(
            f'<div class="ai-risk-card {priority_class}">'
            f'<div class="ai-risk-priority">{html_escape(priority)}</div>'
            f'<strong>{html_escape(risk.get("titulo"))}</strong>'
            f'<p>{html_escape(risk.get("descricao"))}</p>'
            '</div>'
        )

    risks_html = "".join(risks) or (
        '<div class="ai-empty">Não foram identificados riscos críticos ou em atenção '
        'entre as métricas mensuradas.</div>'
    )

    return f"""
    <section class="section ai-section">
        <div class="ai-simple-header">
            <div class="ai-engine-icon">{_genai_engine_icon_svg()}</div>
            <div class="ai-simple-heading">
                <div class="ai-kicker">Análise assistida por inteligência artificial</div>
                <h2 class="section-title ai-title">Resumo Inteligente da Qualidade dos Dados</h2>
                <div class="ai-model-line">{html_escape(model_label)}</div>
            </div>
        </div>

        <div class="ai-summary-card ai-summary-main">
            <h3>Resumo</h3>
            <p>{html_escape(analysis.get('resumo_executivo', 'Sem resumo disponível.'))}</p>
        </div>

        <div class="ai-two-columns">
            <div class="ai-list-card ai-positive-card">
                <h3>Pontos positivos</h3>
                <ul>{positives}</ul>
            </div>
            <div class="ai-list-card ai-recommendation-card">
                <h3>Recomendações</h3>
                <ul>{recommendations}</ul>
            </div>
        </div>

        <h3 class="ai-risk-title">Riscos que precisam de atenção</h3>
        <div class="ai-risk-grid">{risks_html}</div>
    </section>
    """


# =========================================================
# Gráficos
# =========================================================
def fig_to_base64(fig):
    buffer = BytesIO()
    fig.savefig(buffer, format="png", bbox_inches="tight", dpi=160)
    plt.close(fig)
    buffer.seek(0)
    return base64.b64encode(buffer.read()).decode("utf-8")


def chart_matrix(summary_df):
    from matplotlib.ticker import PercentFormatter

    if summary_df.empty:
        return None

    work = summary_df.copy()
    score_col = detect_score_col(work)
    if not score_col:
        return None

    if "row_count" not in work.columns:
        work["row_count"] = range(1, len(work) + 1)

    work["row_count"] = pd.to_numeric(work["row_count"], errors="coerce").fillna(0)
    work["adherence"] = pd.to_numeric(
        work[score_col], errors="coerce"
    ).fillna(0).map(score_to_adherence)

    colors = [chart_color_for_adherence(value) for value in work["adherence"]]
    fig, ax = plt.subplots(figsize=(14, 7.5))
    ax.scatter(work["row_count"], work["adherence"], s=240, alpha=0.88, c=colors)
    for _, row in work.iterrows():
        ax.annotate(
            safe_str(row.get("dataset_label")),
            (row["row_count"], row["adherence"]),
            xytext=(7, 7),
            textcoords="offset points",
            fontsize=10,
        )
    ax.set_title(
        "Matriz Executiva de Aderência de Qualidade",
        fontsize=16, fontweight="bold", pad=16,
    )
    ax.set_xlabel("Quantidade de registros", fontsize=12, labelpad=10)
    ax.set_ylabel("Aderência de qualidade", fontsize=12, labelpad=10)
    ax.set_ylim(0, 105)
    ax.yaxis.set_major_formatter(PercentFormatter(xmax=100, decimals=0))
    ax.tick_params(axis="both", labelsize=11)
    ax.grid(True, alpha=0.3)
    plt.tight_layout(pad=1.5)
    return fig_to_base64(fig)


def chart_distribution(summary_df):
    import textwrap
    from matplotlib.ticker import PercentFormatter

    if summary_df.empty:
        return None

    score_col = detect_score_col(summary_df)
    if not score_col:
        return None

    work = summary_df.copy()
    work["adherence"] = pd.to_numeric(
        work[score_col], errors="coerce"
    ).fillna(0).map(score_to_adherence)
    work = work.sort_values("adherence", ascending=True)

    def short_label(text, width=28, max_lines=3):
        txt = safe_str(text)
        wrapped = textwrap.wrap(txt, width=width)
        if len(wrapped) > max_lines:
            wrapped = wrapped[:max_lines]
            wrapped[-1] = wrapped[-1][: max(0, width - 3)] + "..."
        return "\n".join(wrapped)

    work["dataset_label_plot"] = work["dataset_label"].apply(short_label)
    fig_height = max(6.5, 0.9 * len(work) + 3.0)
    colors = [chart_color_for_adherence(value) for value in work["adherence"]]

    fig, ax = plt.subplots(figsize=(14, fig_height))
    bars = ax.barh(work["dataset_label_plot"], work["adherence"], color=colors)

    ax.set_title(
        "Distribuição da Aderência de Qualidade",
        fontsize=16, fontweight="bold", pad=16,
    )
    ax.set_xlabel("Aderência de qualidade", fontsize=12, labelpad=10)
    ax.set_ylabel("Conjunto de dados", fontsize=12, labelpad=12)
    ax.set_xlim(0, 105)
    ax.xaxis.set_major_formatter(PercentFormatter(xmax=100, decimals=0))
    ax.tick_params(axis="x", labelsize=11)
    ax.tick_params(axis="y", labelsize=11)
    ax.grid(True, axis="x", alpha=0.3)

    for bar, value in zip(bars, work["adherence"]):
        ax.text(
            min(bar.get_width() + 1.0, 101),
            bar.get_y() + bar.get_height() / 2,
            f"{_format_number_pt_br(value, 2)}%",
            va="center",
            fontsize=10.5,
            fontweight="bold",
        )

    plt.tight_layout(pad=1.5)
    return fig_to_base64(fig)


def chart_dimensions(dim_df):
    from matplotlib.ticker import PercentFormatter

    if dim_df.empty:
        return None

    preferred = [
        "dim_acessibilidade", "dim_acuracia", "dim_adaptabilidade", "dim_clareza",
        "dim_comparabilidade", "dim_completude", "dim_confiabilidade",
        "dim_consistencia", "dim_integridade", "dim_rastreabilidade",
        "dim_relevancia", "dim_tempestividade", "dim_unicidade", "dim_validade",
        "dim_integridade_ref", "dim_freshness",
    ]
    dim_cols = [
        column
        for column in preferred
        if column in dim_df.columns
        and pd.to_numeric(dim_df[column], errors="coerce").notna().any()
    ]
    if not dim_cols:
        return None

    averages = {
        COLUMN_LABELS_PT_BR.get(
            column,
            column.replace("dim_", "").replace("_", " ").title(),
        ): score_to_adherence(
            pd.to_numeric(dim_df[column], errors="coerce").dropna().mean()
        )
        for column in dim_cols
    }

    labels = list(averages.keys())
    values = list(averages.values())
    colors = [chart_color_for_adherence(value) for value in values]

    fig, ax = plt.subplots(figsize=(14, 7.5))
    bars = ax.bar(labels, values, color=colors)
    ax.set_title(
        "Aderência Média por Dimensão",
        fontsize=16, fontweight="bold", pad=16,
    )
    ax.set_ylabel("Aderência média", fontsize=12, labelpad=10)
    ax.set_ylim(0, 105)
    ax.yaxis.set_major_formatter(PercentFormatter(xmax=100, decimals=0))
    ax.tick_params(axis="y", labelsize=11)
    ax.tick_params(axis="x", rotation=30, labelsize=10)
    for label in ax.get_xticklabels():
        label.set_horizontalalignment("right")
    ax.grid(True, axis="y", alpha=0.3)

    for bar, value in zip(bars, values):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            min(value + 1.3, 101),
            f"{_format_number_pt_br(value, 1)}%",
            ha="center", va="bottom", fontsize=9.5, fontweight="bold",
        )

    plt.tight_layout(pad=1.6)
    return fig_to_base64(fig)


def chart_top5(top_df):
    from matplotlib.ticker import PercentFormatter

    if top_df.empty or "score" not in top_df.columns:
        return None

    work = top_df.copy()
    work["adherence"] = pd.to_numeric(
        work["score"], errors="coerce"
    ).fillna(0).map(score_to_adherence)
    work = work.sort_values("adherence", ascending=True)
    colors = [chart_color_for_adherence(value) for value in work["adherence"]]

    fig, ax = plt.subplots(figsize=(14, 7))
    bars = ax.barh(work["dataset_label"], work["adherence"], color=colors)
    ax.set_title(
        "Ranking de Aderência de Qualidade — Top 5",
        fontsize=16, fontweight="bold", pad=16,
    )
    ax.set_xlabel("Aderência de qualidade", fontsize=12, labelpad=10)
    ax.set_ylabel("Conjunto de dados", fontsize=12, labelpad=12)
    ax.set_xlim(0, 105)
    ax.xaxis.set_major_formatter(PercentFormatter(xmax=100, decimals=0))
    ax.tick_params(axis="both", labelsize=11)
    ax.grid(True, axis="x", alpha=0.3)

    for bar, value in zip(bars, work["adherence"]):
        ax.text(
            min(bar.get_width() + 1.0, 101),
            bar.get_y() + bar.get_height() / 2,
            f"{_format_number_pt_br(value, 2)}%",
            va="center", fontsize=10.5, fontweight="bold",
        )

    plt.tight_layout(pad=1.5)
    return fig_to_base64(fig)


def chart_history(history_df):
    from matplotlib.ticker import PercentFormatter

    if history_df.empty:
        return None

    work = history_df.copy()
    work["adherence"] = pd.to_numeric(
        work["avg_score"], errors="coerce"
    ).fillna(0).map(score_to_adherence)

    fig, ax = plt.subplots(figsize=(12, 6))
    ax.plot(work["run_id"], work["adherence"], marker="o")
    ax.set_title("Evolução da Aderência de Qualidade")
    ax.set_xlabel("ID da execução")
    ax.set_ylabel("Aderência média")
    ax.set_ylim(0, 105)
    ax.yaxis.set_major_formatter(PercentFormatter(xmax=100, decimals=0))
    ax.tick_params(axis="x", rotation=30)
    ax.grid(True, alpha=0.3)
    return fig_to_base64(fig)


# =========================================================
# Exportações auxiliares
# =========================================================

def export_radar_html(path, run_id, dim_df):
    import plotly.graph_objects as go
    from datetime import datetime

    if dim_df is None or dim_df.empty:
        html = f"""
        <!DOCTYPE html>
        <html lang="pt-BR" translate="no">
        <head>
            <meta charset="utf-8"/>
            <meta name="google" content="notranslate"/>
            <title>Radar de Aderência por Dimensão</title>
            <style>
                body {{ margin:0; font-family: Arial, Helvetica, sans-serif; background:#F5F7FB; color:#1F2937; }}
                .page {{ max-width: 1360px; margin:0 auto; }}
                .cover {{ background: linear-gradient(135deg, #0B1F3A 0%, #15396B 45%, #0F5DFF 100%); color:white; border-radius:0 0 28px 28px; padding:36px 42px 46px 42px; }}
                .section {{ margin:24px; background:#FFFFFF; border:1px solid #C5D0DE; border-radius:24px; padding:24px; }}
                .footer {{ padding:28px; text-align:center; color:#6B7280; font-size:12px; }}
                .cover-kicker {{ text-transform:uppercase; letter-spacing:2px; font-size:12px; opacity:0.85; font-weight:700; }}
                .cover h1 {{ margin:14px 0 10px 0; font-size:38px; line-height:1.1; }}
                .cover p {{ margin:0; font-size:16px; line-height:1.6; color:rgba(255,255,255,0.92); }}
                .cover-meta {{ margin-top:22px; display:flex; gap:12px; flex-wrap:wrap; }}
                .pill {{ background:rgba(255,255,255,0.12); border:1px solid rgba(255,255,255,0.18); color:white; padding:10px 14px; border-radius:999px; font-size:13px; font-weight:700; }}
            </style>
        </head>
        <body>
            <div class="page">
                <section class="cover">
                    <div class="cover-kicker">MJV Data Quality</div>
                    <h1>Radar de Aderência por Dimensão</h1>
                    <p>Visão resumida do percentual de aderência dos dados por dimensão de qualidade.</p>
                    <div class="cover-meta">
                        <span class="pill">Run ID: {html_escape(run_id)}</span>
                        <span class="pill">Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}</span>
                        <span class="pill">Escala: 0% a 100%</span>
                    </div>
                </section>

                <section class="section">
                    <p>Sem dados para radar.</p>
                </section>

                <div class="footer">Aderência por dimensão MJV - 2026</div>
            </div>
        </body>
        </html>
        """
        path.write_text(html, encoding="utf-8")
        print(f"[OK] Radar: {path}")
        return

    work = dim_df.copy()

    if "dataset_label" not in work.columns:
        if "table_name" in work.columns:
            work["dataset_label"] = work["table_name"]
        else:
            work["dataset_label"] = "Dataset"

    dim_map = [
        ("dim_acessibilidade", "Acessibilidade"),
        ("dim_acuracia", "Acurácia"),
        ("dim_adaptabilidade", "Adaptabilidade"),
        ("dim_clareza", "Clareza"),
        ("dim_comparabilidade", "Comparabilidade"),
        ("dim_completude", "Completude"),
        ("dim_confiabilidade", "Confiabilidade"),
        ("dim_consistencia", "Consistência"),
        ("dim_integridade", "Integridade"),
        ("dim_rastreabilidade", "Rastreabilidade"),
        ("dim_relevancia", "Relevância"),
        ("dim_tempestividade", "Tempestividade"),
        ("dim_integridade_ref", "Integridade"),
        ("dim_freshness", "Atualidade"),
        ("dim_validade", "Validade"),
        ("dim_unicidade", "Unicidade"),
    ]

    available_dims = [
        (col, label)
        for col, label in dim_map
        if col in work.columns
        and pd.to_numeric(work[col], errors="coerce").notna().any()
    ]

    if not available_dims:
        html = f"""
        <!DOCTYPE html>
        <html lang="pt-BR" translate="no">
        <head>
            <meta charset="utf-8"/>
            <meta name="google" content="notranslate"/>
            <title>Radar de Aderência por Dimensão</title>
        </head>
        <body>
            <h1>Radar de Aderência por Dimensão</h1>
            <p>Sem colunas de dimensão disponíveis.</p>
        </body>
        </html>
        """
        path.write_text(html, encoding="utf-8")
        print(f"[OK] Radar: {path}")
        return

    categories = [label for _, label in available_dims]
    categories_closed = categories + [categories[0]]

    fig = go.Figure()
    dataset_labels = []

    for _, row in work.iterrows():
        dataset = safe_str(row.get("dataset_label")) or "Dataset não identificado"
        dataset_labels.append(dataset)

        values = []
        hover_parts = []
        for col, label in available_dims:
            raw_value = pd.to_numeric(
                pd.Series([row.get(col)]), errors="coerce"
            ).iloc[0]
            if pd.isna(raw_value):
                values.append(None)
                hover_parts.append(f"{label}: N/A")
            else:
                adherence = score_to_adherence(raw_value)
                values.append(adherence)
                hover_parts.append(
                    f"{label}: {_format_number_pt_br(adherence, 2)}%"
                )

        values_closed = values + [values[0]]
        hover_lines = "<br>".join(hover_parts)

        fig.add_trace(
            go.Scatterpolar(
                r=values_closed,
                theta=categories_closed,
                fill="toself",
                name=dataset,
                visible=True,
                hovertemplate=f"<b>{dataset}</b><br>{hover_lines}<extra></extra>",
            )
        )

    buttons = [
        dict(
            label="Todos os datasets",
            method="update",
            args=[
                {"visible": [True] * len(dataset_labels)},
                {"title": "Radar de Aderência por Dimensão - Todos os conjuntos de dados"},
            ],
        )
    ]

    for i, dataset in enumerate(dataset_labels):
        visible = [False] * len(dataset_labels)
        visible[i] = True
        buttons.append(
            dict(
                label=dataset[:60],
                method="update",
                args=[
                    {"visible": visible},
                    {"title": f"Radar de Aderência por Dimensão - {dataset}"},
                ],
            )
        )

    fig.update_layout(
        title="Radar de Aderência por Dimensão - Todos os conjuntos de dados",
        polar=dict(
            radialaxis=dict(
                visible=True,
                range=[0, 100],
                tick0=0,
                dtick=20,
            )
        ),
        template="plotly_white",
        height=760,
        updatemenus=[
            dict(
                type="dropdown",
                direction="down",
                x=1.02,
                y=1.12,
                xanchor="left",
                yanchor="top",
                buttons=buttons,
                showactive=True,
            )
        ],
        legend=dict(
            orientation="v",
            yanchor="top",
            y=1.0,
            xanchor="left",
            x=1.02
        ),
        margin=dict(l=60, r=280, t=90, b=50),
    )

    chart_html = fig.to_html(
        include_plotlyjs="cdn",
        full_html=False,
        config={
            "displaylogo": False,
            "scrollZoom": True,
            "responsive": True,
        },
    )

    html = f"""
    <!DOCTYPE html>
    <html lang="pt-BR" translate="no">
    <head>
        <meta charset="utf-8"/>
            <meta name="google" content="notranslate"/>
        <title>Radar de Aderência por Dimensão</title>
        <style>
            body {{ margin:0; font-family: Arial, Helvetica, sans-serif; background:#F5F7FB; color:#1F2937; }}
            .page {{ max-width: 1360px; margin:0 auto; }}
            .cover {{ background: linear-gradient(135deg, #0B1F3A 0%, #15396B 45%, #0F5DFF 100%); color:white; border-radius:0 0 28px 28px; padding:36px 42px 46px 42px; box-shadow:0 18px 50px rgba(11,31,58,0.18); }}
            .cover-kicker {{ text-transform:uppercase; letter-spacing:2px; font-size:12px; opacity:0.85; font-weight:700; }}
            .cover h1 {{ margin:14px 0 10px 0; font-size:38px; line-height:1.1; }}
            .cover p {{ margin:0; max-width:880px; font-size:16px; line-height:1.6; color:rgba(255,255,255,0.92); }}
            .cover-meta {{ margin-top:22px; display:flex; gap:12px; flex-wrap:wrap; }}
            .pill {{ background:rgba(255,255,255,0.12); border:1px solid rgba(255,255,255,0.18); color:white; padding:10px 14px; border-radius:999px; font-size:13px; font-weight:700; }}
            .section {{ margin:24px; background:#FFFFFF; border:1px solid #C5D0DE; border-radius:24px; padding:24px; box-shadow:0 8px 24px rgba(15,23,42,0.05); }}
            .footer {{ padding:28px; text-align:center; color:#6B7280; font-size:12px; }}
            .hint {{ font-size:13px; color:#6B7280; margin-top:8px; }}
        </style>
    </head>
    <body>
        <div class="page">
            <section class="cover">
                <div class="cover-kicker">MJV Data Quality</div>
                <h1>Radar de Aderência por Dimensão</h1>
                <p>Visão resumida do percentual de aderência dos dados por dimensão de qualidade.</p>
                <div class="cover-meta">
                    <span class="pill">Run ID: {html_escape(run_id)}</span>
                    <span class="pill">Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}</span>
                    <span class="pill">Escala: 0% a 100%</span>
                </div>
            </section>

            <section class="section">
                {chart_html}
                <div class="hint">Nota: Use o menu no canto superior direito para filtrar por dataset.</div>
            </section>

            <section class="section">
                <h2 style="margin:0 0 18px 0; font-size:28px; color:#0B1F3A;">Detalhamento das Métricas Avaliadas</h2>
                <p style="margin:0 0 18px 0; font-size:15px; line-height:1.7; color:#4B5563;">
                    Abaixo está o significado de cada dimensão monitorada no radar de qualidade MJV.
                </p>
                <div style="overflow-x:auto;">
                    <table style="width:100%; border-collapse:separate; border-spacing:0; font-size:15px; line-height:1.6; border:1px solid #B8C5D5; border-radius:18px; overflow:hidden;">
                        <thead>
                            <tr style="background:linear-gradient(135deg, #0B1F3A 0%, #15396B 100%); color:#FFFFFF;">
                                <th style="text-align:left; padding:16px 18px; width:220px; font-size:14px; letter-spacing:0.3px;">Métrica</th>
                                <th style="text-align:left; padding:16px 18px; font-size:14px; letter-spacing:0.3px;">Detalhamento</th>
                            </tr>
                        </thead>
                        <tbody>
                            {bacen_dimension_description_rows()}
                        </tbody>
                    </table>
                </div>
                <div class="hint">Leitura recomendada: percentuais mais baixos indicam maior prioridade de correção naquela dimensão específica.</div>
            </section>


            <div class="footer">Aderência por dimensão MJV - 2026</div>
        </div>
    </body>
    </html>
    """
    path.write_text(html, encoding="utf-8")
    print(f"[OK] Radar: {path}")


def export_history_html(path, history_df):
    import plotly.graph_objects as go
    from datetime import datetime

    if history_df is None or history_df.empty:
        html = f"""
        <!DOCTYPE html>
        <html lang="pt-BR" translate="no">
        <head>
            <meta charset="utf-8"/>
            <meta name="google" content="notranslate"/>
            <style>
                body {{ margin:0; font-family:Arial, Helvetica, sans-serif; background:#F5F7FB; color:#1F2937; }}
                .page {{ max-width:1200px; margin:0 auto; }}
                .cover {{ background:linear-gradient(135deg,#0B1F3A,#0F5DFF); color:white; padding:30px; border-radius:0 0 24px 24px; }}
                .section {{ margin:20px; background:white; padding:20px; border:1px solid #C5D0DE; border-radius:16px; box-shadow:0 6px 20px rgba(15,23,42,0.06); }}
                .footer {{ text-align:center; padding:20px; font-size:12px; color:#6B7280; }}
            </style>
        </head>
        <body>
            <div class="page">
                <div class="cover">
                    <h1>Evolução da Aderência de Qualidade</h1>
                    <p>Histórico do percentual de aderência ao longo das execuções.</p>
                </div>
                <div class="section"><p>Sem dados históricos disponíveis.</p></div>
                <div class="footer">Histórico de aderência MJV - 2026</div>
            </div>
        </body>
        </html>
        """
        path.write_text(html, encoding="utf-8")
        return

    work = history_df.sort_values("run_id").copy()
    score_series = (
        work["avg_score"] if "avg_score" in work.columns else work["score"]
    )
    work["adherence"] = pd.to_numeric(
        score_series, errors="coerce"
    ).fillna(0).map(score_to_adherence)

    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=work["run_id"],
            y=work["adherence"],
            mode="lines+markers",
            name="Aderência de qualidade",
            hovertemplate=(
                "<b>Execução:</b> %{x}<br>"
                "<b>Aderência:</b> %{y:.2f}%<extra></extra>"
            ),
        )
    )
    fig.update_layout(
        title="Evolução da Aderência de Qualidade",
        xaxis_title="ID da execução",
        yaxis_title="Aderência de qualidade",
        yaxis=dict(range=[0, 100], ticksuffix="%", dtick=20),
        template="plotly_white",
        height=500,
    )

    chart_html = fig.to_html(
        include_plotlyjs="cdn",
        full_html=False,
        config={"displaylogo": False, "responsive": True},
    )

    html = f"""
    <!DOCTYPE html>
    <html lang="pt-BR" translate="no">
    <head>
        <meta charset="utf-8"/>
            <meta name="google" content="notranslate"/>
        <style>
            body {{ margin:0; font-family:Arial, Helvetica, sans-serif; background:#F5F7FB; color:#1F2937; }}
            .page {{ max-width:1360px; margin:0 auto; }}
            .cover {{ background:linear-gradient(135deg,#0B1F3A 0%,#15396B 45%,#0F5DFF 100%); color:white; padding:36px; border-radius:0 0 28px 28px; }}
            .cover h1 {{ margin:0; font-size:32px; }}
            .cover p {{ margin-top:10px; }}
            .meta {{ margin-top:15px; font-size:13px; opacity:0.9; }}
            .section {{ margin:24px; background:#FFFFFF; border:1px solid #C5D0DE; border-radius:20px; padding:24px; box-shadow:0 6px 20px rgba(15,23,42,0.06); }}
            .footer {{ padding:28px; text-align:center; color:#6B7280; font-size:12px; }}
        </style>
    </head>
    <body>
        <div class="page">
            <div class="cover">
                <h1>Evolução da Aderência de Qualidade</h1>
                <p>Histórico do percentual de aderência dos dados ao longo das execuções do Data Quality MJV.</p>
                <div class="meta">Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}</div>
            </div>
            <div class="section">{chart_html}</div>
            <div class="footer">Histórico de aderência MJV - 2026</div>
        </div>
    </body>
    </html>
    """

    path.write_text(html, encoding="utf-8")
    print(f"[OK] Histórico: {path}")


def export_support_files(outdir, run_id, detail_df, dim_df, history_df, attention_df, bacen_detail_df=None):
    detail_df.to_csv(outdir / "dq_current_detail.csv", index=False, encoding="utf-8-sig")
    dim_df.to_csv(outdir / "dq_dimension_scores_current.csv", index=False, encoding="utf-8-sig")
    if "score_bacen" in dim_df.columns:
        dim_df.to_csv(outdir / "dq_bacen_summary_current.csv", index=False, encoding="utf-8-sig")
    if bacen_detail_df is not None and not bacen_detail_df.empty:
        bacen_detail_df.to_csv(outdir / "dq_bacen_dimension_detail_current.csv", index=False, encoding="utf-8-sig")

    recs = []
    if attention_df.empty:
        recs.append({"run_id": run_id, "prioridade": "Média", "dataset_label": "", "recomendacao": "Nenhuma recomendação disponível para o run atual."})
    else:
        for _, row in attention_df.head(10).iterrows():
            recs.append({
                "run_id": run_id,
                "prioridade": "Alta",
                "dataset_label": safe_str(row.get("dataset_label")),
                "recomendacao": f"Revisar a coluna {safe_str(row.get('column_name'))} do dataset {safe_str(row.get('dataset_label'))}."
            })
    pd.DataFrame(recs).to_csv(outdir / "dq_ai_recommendations_current.csv", index=False, encoding="utf-8-sig")

    export_radar_html(outdir / "dq_radar_chart.html", run_id, dim_df)
    export_history_html(outdir / "dq_history_chart.html", history_df)


# =========================================================
# Excel e HTML
# =========================================================
def export_excel(path, kpis, summary_df, dim_df, detail_df, top_df, bottom_df, attention_df, history_df, bacen_detail_df=None, bacen_catalog_df=None, semantic_df=None, semantic_invalid_df=None, genai_result=None):
    """Gera Excel executivo com aderência percentual e detalhe técnico preservado."""
    kpi_df = pd.DataFrame([
        {
            "Fontes avaliadas": kpis["datasets"],
            "Aderência de qualidade": kpis["aderencia_percentual"] / 100.0,
            "Nota técnica equivalente": kpis["score_medio"],
            "Classificação geral": kpis["classificacao"],
            "Melhor fonte": kpis["melhor_dataset"],
            "Melhor aderência": kpis["melhor_aderencia"] / 100.0,
            "Pior fonte": kpis["pior_dataset"],
            "Pior aderência": kpis["pior_aderencia"] / 100.0,
            "Fontes críticas": kpis["criticos"],
            "Fontes em atenção": kpis["atencao"],
            "Fontes satisfatórias": kpis["satisfatorios"],
        }
    ])

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        kpi_df.to_excel(writer, sheet_name="Resumo", index=False)
        prepare_dataframe_for_excel(
            summary_df, "summary"
        ).to_excel(writer, sheet_name="Aderência por Fonte", index=False)
        prepare_dataframe_for_excel(
            dim_df, "dimensions"
        ).to_excel(writer, sheet_name="Aderência por Dimensão", index=False)

        if bacen_catalog_df is not None:
            prepare_dataframe_for_excel(
                bacen_catalog_df, "bacen_catalog"
            ).to_excel(writer, sheet_name="Visão Dimensões BACEN", index=False)

        if bacen_detail_df is not None:
            prepare_dataframe_for_excel(
                bacen_detail_df, "bacen_detail"
            ).to_excel(writer, sheet_name="Controles BACEN", index=False)

        if semantic_df is not None and not semantic_df.empty:
            prepare_dataframe_for_excel(
                semantic_df, "semantic"
            ).to_excel(writer, sheet_name="Validações Brasil", index=False)

        if semantic_invalid_df is not None and not semantic_invalid_df.empty:
            prepare_dataframe_for_excel(
                semantic_invalid_df, "semantic_invalid"
            ).to_excel(writer, sheet_name="Amostras Inválidas", index=False)

        prepare_dataframe_for_excel(
            top_df, "top_bottom"
        ).to_excel(writer, sheet_name="Top 5", index=False)
        prepare_dataframe_for_excel(
            bottom_df, "top_bottom"
        ).to_excel(writer, sheet_name="Bottom 5", index=False)
        prepare_dataframe_for_excel(
            attention_df, "attention"
        ).to_excel(writer, sheet_name="Maior Atenção", index=False)
        prepare_dataframe_for_excel(
            history_df, "history"
        ).to_excel(writer, sheet_name="Histórico", index=False)
        prepare_dataframe_for_excel(
            detail_df, "technical"
        ).to_excel(writer, sheet_name="Detalhe Técnico", index=False)
        interpretation_criteria_dataframe().to_excel(
            writer, sheet_name="Critérios de Leitura", index=False
        )
        genai_analysis_dataframe(genai_result).to_excel(
            writer, sheet_name="Análise GenAI", index=False
        )

        _style_excel_workbook(writer)

    print(f"[OK] Excel: {path}")


def export_html(path, run_id, kpis, summary_df, dim_df, detail_df, top_df, bottom_df, attention_df, history_df, bacen_detail_df=None, bacen_catalog_df=None, semantic_df=None, semantic_invalid_df=None, genai_result=None):
    matrix_img = chart_matrix(summary_df)
    dist_img = chart_distribution(summary_df)
    dim_img = chart_dimensions(dim_df)
    top_img = chart_top5(top_df)

    chart_block = ""
    if matrix_img:
        chart_block += f"""
        <section class="section">
            <h2 class="section-title">Matriz Executiva de Aderência de Qualidade</h2>
            <img class="chart" src="data:image/png;base64,{matrix_img}" alt="Matriz de Aderência de Qualidade" />
        </section>
        """

    panel_imgs = []
    if dist_img:
        panel_imgs.append(f'<img class="chart" src="data:image/png;base64,{dist_img}" alt="Distribuição da Aderência de Qualidade" />')
    if dim_img:
        panel_imgs.append(f'<img class="chart" src="data:image/png;base64,{dim_img}" alt="Aderência por Dimensão" />')

    panel_block = ""
    if panel_imgs:
        panel_block = f"""
        <section class="section">
            <h2 class="section-title">Painel Analítico</h2>
            <div class="chart-grid">
                {''.join(panel_imgs)}
            </div>
        </section>
        """

    rank_block = ""
    if top_img:
        rank_block = f"""
        <section class="section">
            <h2 class="section-title">Ranking de Aderência de Qualidade — Top 5</h2>
            <div class="chart-grid-single">
                <img class="chart" src="data:image/png;base64,{top_img}" alt="Ranking Top 5" />
            </div>
        </section>
        """

    html = f"""<!DOCTYPE html>
<html lang="pt-BR" translate="no">
<head>
<meta charset="UTF-8" />
<meta name="google" content="notranslate" />
<meta name="viewport" content="width=device-width, initial-scale=1.0" />
<title>Relatório de Validação e Qualidade de Dados</title>
<style>
    * {{ box-sizing: border-box; }}
    body {{ margin:0; font-family: Arial, Helvetica, sans-serif; background:#F5F7FB; color:#1F2937; }}
    .page {{ max-width: 1320px; margin: 0 auto; }}
    .cover {{ background: linear-gradient(135deg, #0B1F3A 0%, #15396B 45%, #0F5DFF 100%); color:white; border-radius:0 0 28px 28px; padding:36px 42px 46px 42px; box-shadow:0 18px 50px rgba(11,31,58,0.18); }}
    .cover-top {{ display:flex; justify-content:space-between; align-items:center; gap:20px; }}
    .cover-kicker {{ text-transform:uppercase; letter-spacing:2px; font-size:12px; opacity:0.85; font-weight:700; }}
    .cover h1 {{ margin:14px 0 10px 0; font-size:42px; line-height:1.08; }}
    .cover p {{ margin:0; max-width:880px; font-size:17px; line-height:1.6; color:rgba(255,255,255,0.92); }}
    .cover-meta {{ margin-top:22px; display:flex; gap:12px; flex-wrap:wrap; }}
    .pill {{ background:rgba(255,255,255,0.12); border:1px solid rgba(255,255,255,0.18); color:white; padding:10px 14px; border-radius:999px; font-size:13px; font-weight:700; }}
    .container {{ padding:28px; }}
    .section {{ margin-top:24px; background:#FFFFFF; border:1px solid #C5D0DE; border-radius:24px; padding:26px; box-shadow:0 8px 24px rgba(15,23,42,0.07); }}
    .section-title {{ margin:0 0 18px 0; color:#0B1F3A; font-size:28px; }}
    .section-subtitle {{ margin:0 0 8px 0; color:#0B1F3A; font-size:20px; }}
    .lead {{ color:#6B7280; line-height:1.7; font-size:16px; }}
    .metrics-grid {{ display:grid; grid-template-columns:repeat(5, minmax(0,1fr)); gap:16px; }}
    .metric-card {{ background:white; border:1px solid #C5D0DE; border-radius:20px; padding:18px; min-height:132px; box-shadow:0 3px 10px rgba(15,23,42,0.05); }}
    .metric-title {{ color:#6B7280; font-size:13px; font-weight:700; text-transform:uppercase; letter-spacing:.5px; }}
    .metric-value {{ color:#0B1F3A; font-size:32px; font-weight:800; margin-top:8px; }}
    .metric-value-centered {{ text-align:center; width:100%; }}
    .metric-value-classification {{ text-transform:none; }}
    .metric-subtitle {{ color:#6B7280; font-size:13px; margin-top:8px; line-height:1.5; }}
    .split {{ display:grid; grid-template-columns:1.15fr .85fr; gap:18px; align-items:start; }}
    .panel {{ background:white; border:1px solid #C5D0DE; border-radius:20px; padding:18px; box-shadow:0 3px 10px rgba(15,23,42,0.05); }}
    .chart {{ width:100%; border-radius:16px; border:1px solid #C1CCDA; background:white; box-shadow:0 3px 10px rgba(15,23,42,0.04); }}
    .chart-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:18px; }}
    .chart-grid-single {{ display:grid; grid-template-columns:1fr; gap:18px; }}
    .callout {{ background:linear-gradient(135deg, #EAF2FF 0%, white 100%); border:1px solid #B8C9DF; border-radius:20px; padding:20px; box-shadow:0 3px 10px rgba(15,23,42,0.05); }}
    ul.exec {{ margin:12px 0 0 0; padding-left:22px; }}
    ul.exec li {{ margin-bottom:10px; line-height:1.6; }}
    .table-toolbar {{ display:flex; justify-content:space-between; gap:16px; align-items:center; margin:0 0 10px 0; color:#64748B; font-size:12px; }}
    .table-count {{ font-weight:700; color:#334155; }}
    .table-hint {{ text-align:right; }}
    .report-table-wrapper {{ width:100%; overflow:auto; max-height:620px; border:1px solid #B8C5D5; border-radius:16px; background:#FFFFFF; box-shadow:0 3px 10px rgba(15,23,42,0.04); }}
    .report-table {{ width:max-content; min-width:100%; border-collapse:separate; border-spacing:0; font-size:13px; }}
    .report-table th {{ position:sticky; top:0; z-index:4; text-align:left; background:#0B1F3A; color:#FFFFFF; padding:12px 14px; min-width:130px; max-width:260px; border-right:1px solid rgba(255,255,255,0.12); white-space:normal; line-height:1.35; overflow-wrap:anywhere; word-break:break-word; }}
    .report-table td {{ padding:11px 14px; border-bottom:1px solid #E5E7EB; border-right:1px solid #EEF2F7; vertical-align:top; background:#FFFFFF; min-width:130px; max-width:360px; line-height:1.45; white-space:normal; overflow-wrap:anywhere; word-break:break-word; }}
    .report-table tbody tr:nth-child(even) td {{ background:#F8FAFC; }}
    .report-table tbody tr:hover td {{ background:#EFF6FF; }}
    .report-table th:first-child {{ left:0; z-index:6; min-width:210px; width:210px; max-width:210px; }}
    .report-table td:first-child {{ position:sticky; left:0; z-index:2; min-width:210px; width:210px; max-width:210px; font-weight:700; color:#0F172A; box-shadow:3px 0 6px rgba(15,23,42,0.05); overflow-wrap:anywhere; }}
    .report-table tbody tr:nth-child(odd) td:first-child {{ background:#FFFFFF; }}
    .report-table tbody tr:nth-child(even) td:first-child {{ background:#F8FAFC; }}
    .report-table tbody tr:hover td:first-child {{ background:#EFF6FF; }}
    .status-badge {{ display:inline-flex; align-items:center; justify-content:center; padding:5px 10px; border-radius:999px; font-size:12px; font-weight:800; white-space:nowrap; }}
    .status-excellent {{ background:#E7F5F0; color:#0F7B5B; }}
    .status-good {{ background:#E8F7EF; color:#2EAD67; }}
    .status-adequate {{ background:#EAF2FF; color:#0F5DFF; }}
    .status-warning {{ background:#FFF4E8; color:#F28C28; }}
    .status-critical {{ background:#FDECEC; color:#D64545; }}
    .status-neutral {{ background:#F1F5F9; color:#475569; }}
    td.cell-score-excellent {{ background:#E7F5F0 !important; color:#0F7B5B; font-weight:800; text-align:right; }}
    td.cell-score-good {{ background:#E8F7EF !important; color:#2EAD67; font-weight:800; text-align:right; }}
    td.cell-score-adequate {{ background:#EAF2FF !important; color:#0F5DFF; font-weight:800; text-align:right; }}
    td.cell-score-warning {{ background:#FFF4E8 !important; color:#F28C28; font-weight:800; text-align:right; }}
    td.cell-score-critical {{ background:#FDECEC !important; color:#D64545; font-weight:800; text-align:right; }}
    td.cell-score-na {{ background:#F1F5F9 !important; color:#64748B; font-weight:800; text-align:center; }}
    td.cell-contribution {{ background:#F8FAFC !important; color:#334155; font-weight:800; text-align:right; }}
    td.cell-long-text {{ min-width:280px; width:360px; max-width:460px; color:#475569; white-space:normal; overflow-wrap:anywhere; word-break:break-word; }}
    th.cell-column-name, td.cell-column-name {{ min-width:300px !important; width:300px !important; max-width:300px !important; white-space:normal !important; overflow-wrap:anywhere !important; word-break:break-word !important; }}
    th.cell-row-number, td.cell-row-number {{ min-width:110px !important; width:110px !important; max-width:110px !important; text-align:center; white-space:nowrap !important; }}
    th.cell-masked-value, td.cell-masked-value,
    th.cell-sample-value, td.cell-sample-value {{ min-width:250px !important; width:250px !important; max-width:250px !important; white-space:normal !important; overflow-wrap:anywhere !important; word-break:break-all !important; font-family:Consolas, "Courier New", monospace; }}
    th.cell-records-by-type, td.cell-records-by-type {{ min-width:320px !important; width:320px !important; max-width:320px !important; white-space:normal !important; overflow-wrap:anywhere !important; }}
    th.cell-data-type, td.cell-data-type {{ min-width:105px !important; width:105px !important; max-width:105px !important; white-space:normal !important; }}
    th.cell-object-name, td.cell-object-name,
    th.cell-dataset-name, td.cell-dataset-name {{ min-width:210px !important; width:210px !important; max-width:210px !important; white-space:normal !important; overflow-wrap:anywhere !important; }}
    .table-empty {{ padding:24px; border:1px dashed #CBD5E1; border-radius:14px; background:#F8FAFC; color:#64748B; text-align:center; }}

    .ai-section {{ border:1px solid #9FB7D7; background:linear-gradient(180deg,#F7FAFF 0%,#FFFFFF 100%); }}
    .ai-simple-header {{ display:flex; align-items:center; gap:16px; margin-bottom:18px; }}
    .ai-engine-icon {{ width:64px; height:64px; flex:0 0 64px; filter:drop-shadow(0 6px 12px rgba(15,93,255,0.18)); }}
    .ai-engine-icon svg {{ width:100%; height:100%; display:block; }}
    .ai-simple-heading {{ min-width:0; }}
    .ai-kicker {{ color:#0F5DFF; font-size:12px; font-weight:800; letter-spacing:1px; text-transform:uppercase; margin-bottom:5px; }}
    .ai-title {{ margin:0 0 6px 0; }}
    .ai-model-line {{ color:#64748B; font-size:12px; font-weight:700; overflow-wrap:anywhere; }}
    .ai-summary-card {{ padding:18px 20px; border:1px solid #C5D0DE; border-radius:16px; background:#FFFFFF; }}
    .ai-summary-main {{ border-left:5px solid #0F5DFF; }}
    .ai-summary-card h3, .ai-list-card h3 {{ margin:0 0 10px 0; color:#0B1F3A; font-size:18px; }}
    .ai-summary-card p {{ margin:0; color:#334155; font-size:16px; line-height:1.75; }}
    .ai-two-columns {{ display:grid; grid-template-columns:1fr 1fr; gap:14px; margin-top:14px; }}
    .ai-list-card {{ padding:18px 20px; border:1px solid #C5D0DE; border-radius:16px; background:#FFFFFF; }}
    .ai-positive-card {{ border-top:4px solid #0F7B5B; }}
    .ai-recommendation-card {{ border-top:4px solid #0F5DFF; }}
    .ai-list-card ul {{ margin:0; padding-left:20px; color:#475569; line-height:1.7; }}
    .ai-list-card li {{ margin-bottom:9px; }}
    .ai-risk-title {{ color:#0B1F3A; margin:22px 0 12px 0; font-size:18px; }}
    .ai-risk-grid {{ display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:12px; }}
    .ai-risk-card {{ position:relative; padding:17px; border:1px solid #E3B7B7; border-radius:14px; background:#FFF9F9; color:#475569; }}
    .ai-risk-card strong {{ display:block; color:#7F1D1D; margin:7px 0 8px 0; font-size:15px; }}
    .ai-risk-card p {{ margin:0; line-height:1.6; }}
    .ai-risk-priority {{ display:inline-flex; padding:4px 9px; border-radius:999px; font-size:11px; font-weight:800; text-transform:uppercase; }}
    .ai-risk-high .ai-risk-priority {{ background:#FDECEC; color:#D64545; }}
    .ai-risk-medium {{ border-color:#F5C889; background:#FFF9F0; }}
    .ai-risk-medium .ai-risk-priority {{ background:#FFF4E8; color:#F28C28; }}
    .ai-risk-low {{ border-color:#B8C9DF; background:#F8FAFC; }}
    .ai-risk-low .ai-risk-priority {{ background:#EAF2FF; color:#0F5DFF; }}
    .ai-empty {{ grid-column:1 / -1; padding:17px; border:1px dashed #CBD5E1; border-radius:12px; color:#64748B; background:#F8FAFC; text-align:center; }}

    .glossary-block {{ margin-top:18px; padding:18px; border:1px solid #C5D0DE; border-radius:16px; background:#F8FAFC; }}
    .glossary-block h3 {{ margin:0 0 12px 0; color:#0B1F3A; font-size:18px; }}
    .glossary-grid {{ display:grid; grid-template-columns:repeat(3, minmax(0,1fr)); gap:12px; }}
    .glossary-card {{ display:flex; flex-direction:column; gap:6px; padding:14px; border:1px solid #D9E1EC; border-radius:12px; background:#FFFFFF; color:#475569; line-height:1.5; }}
    .glossary-card strong {{ color:#0B1F3A; }}
    .classification-title {{ margin-top:20px !important; }}
    .classification-legend {{ display:grid; grid-template-columns:auto 1fr; gap:9px 12px; align-items:center; color:#475569; font-size:13px; }}
    .range-note {{ margin:0 0 16px 0; padding:16px 18px; border:1px solid #B8C9DF; border-radius:14px; background:#F8FAFC; color:#475569; line-height:1.65; }}
    .range-table-wrapper {{ max-height:none; }}
    .range-table th:first-child, .range-table td:first-child {{ min-width:180px; }}
    table {{ width:100%; border-collapse:collapse; font-size:14px; }}
    th {{ text-align:left; background:#0B1F3A; color:white; padding:12px 14px; }}
    td {{ padding:12px 14px; border-bottom:1px solid #D9E1EC; vertical-align:top; }}
    tr:nth-child(even) td {{ background:#FBFCFE; }}
    .footer {{ padding:28px; text-align:center; color:#6B7280; font-size:12px; }}
    @media (max-width: 1100px) {{ .metrics-grid {{ grid-template-columns:repeat(2, minmax(0,1fr)); }} .split, .chart-grid, .glossary-grid, .ai-two-columns, .ai-risk-grid {{ grid-template-columns:1fr; }} }}
    @media (max-width: 720px) {{ .metrics-grid {{ grid-template-columns:1fr; }} .container {{ padding:16px; }} .cover {{ padding:26px 22px 34px 22px; border-radius:0 0 20px 20px; }} .cover h1 {{ font-size:30px; }} }}
</style>
</head>
<body class="notranslate">
    <div class="page">
        <section class="cover">
            <div class="cover-top">
                <div>
                    <div class="cover-kicker">MJV Data Quality</div>
                    <h1>Relatório de Validação e Qualidade de Dados</h1>
                    <p>Visão gerencial consolidada do percentual de aderência das fontes avaliadas, com leitura executiva, matriz de qualidade, ranking de desempenho e direcionadores de priorização para tomada de decisão.</p>
                    <div class="cover-meta">
                        <span class="pill">Run ID: {html_escape(run_id)}</span>
                        <span class="pill">Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}</span>
                        <span class="pill">Classificação Geral: {html_escape(normalize_status_label(kpis['classificacao']))}</span>
                        <span class="pill">Aderência: {_format_number_pt_br(kpis['aderencia_percentual'], 2)}%</span>
                    </div>
                </div>
                <div><div style="font-size:28px;font-weight:800;letter-spacing:1px;">MJV</div></div>
            </div>
        </section>

        <div class="container">
            <section class="section">
                <h2 class="section-title">Sumário Executivo</h2>
                <div class="metrics-grid">
                    <div class="metric-card metric-card-sources" style="border-top:5px solid #0F5DFF;">
                        <div class="metric-title">Fontes Avaliadas</div>
                        <div class="metric-value metric-value-centered">{kpis['datasets']}</div>
                        <div class="metric-subtitle">Escopo total da avaliação executada.</div>
                    </div>
                    <div class="metric-card" style="border-top:5px solid {status_color(kpis['classificacao'])};">
                        <div class="metric-title">Aderência de Qualidade</div>
                        <div class="metric-value">{_format_number_pt_br(kpis['aderencia_percentual'], 2)}%</div>
                        <div class="metric-subtitle">Nota técnica equivalente: {_format_number_pt_br(kpis['score_medio'], 2)} de 10.</div>
                    </div>
                    <div class="metric-card" style="border-top:5px solid {status_color(kpis['classificacao'])};">
                        <div class="metric-title">Classificação Geral</div>
                        <div class="metric-value metric-value-classification">{html_escape(normalize_status_label(kpis['classificacao']))}</div>
                        <div class="metric-subtitle">Leitura executiva da maturidade atual.</div>
                    </div>
                    <div class="metric-card" style="border-top:5px solid #0F7B5B;">
                        <div class="metric-title">Melhor Fonte</div>
                        <div class="metric-value">{_format_number_pt_br(kpis['melhor_aderencia'], 2)}%</div>
                        <div class="metric-subtitle">{html_escape(kpis['melhor_dataset'])} • nota {_format_number_pt_br(kpis['melhor_score'], 2)}</div>
                    </div>
                    <div class="metric-card" style="border-top:5px solid #D64545;">
                        <div class="metric-title">Pior Fonte</div>
                        <div class="metric-value">{_format_number_pt_br(kpis['pior_aderencia'], 2)}%</div>
                        <div class="metric-subtitle">{html_escape(kpis['pior_dataset'])} • nota {_format_number_pt_br(kpis['pior_score'], 2)}</div>
                    </div>
                </div>
            </section>

            {genai_analysis_html(genai_result)}

            <section class="section">
                <div class="split">
                    <div class="panel">
                        <h3 class="section-subtitle">Resumo de Avaliação</h3>
                        <p class="lead">{html_escape(summary_message(kpis))}</p>
                        <div style="margin-top:14px;"><span style="display:inline-block;padding:8px 14px;border-radius:999px;background:{status_color(kpis['classificacao'])};color:#FFFFFF;font-weight:700;font-size:13px;">{html_escape(normalize_status_label(kpis['classificacao']))}</span></div>
                    </div>
                    <div class="callout">
                        <h3 class="section-subtitle" style="margin-top:0;">Resumo das Fontes de Dados</h3>
                        {executive_list_html(kpis)}
                    </div>
                </div>
            </section>

            {chart_block}
            {panel_block}
            {rank_block}

            <section class="section">
                <h2 class="section-title">Resumo por Conjunto de Dados</h2>
                <p class="lead">A distribuição por tipo é calculada a partir da coluna de tipo de registro. Quando o arquivo não possui essa informação, o relatório apresenta N/A.</p>
                {df_to_html(summary_df, "summary")}
            </section>

            <section class="section">
                <h2 class="section-title">Aderência por Dimensão</h2>
                {df_to_html(dim_df, "dimensions")}
            </section>

            <section class="section">
                <h2 class="section-title">Visão Consolidada das Dimensões BACEN</h2>
                <p class="lead">Síntese das dimensões regulatórias, com média, amplitude dos resultados, cobertura, peso, tipo de avaliação e classificação.</p>
                {df_to_html(bacen_catalog_df if bacen_catalog_df is not None else pd.DataFrame(), "bacen_catalog")}
            </section>

            <section class="section">
                <h2 class="section-title">Controles, Pesos e Evidências BACEN</h2>
                <p class="lead">Detalhamento por conjunto de dados e dimensão. Itens sem evidência são apresentados como “Não avaliado”, sem serem convertidos automaticamente em aderência zero. A contribuição ponderada representa quanto cada dimensão adiciona ao índice final, em pontos percentuais.</p>
                {df_to_html(bacen_detail_df if bacen_detail_df is not None else pd.DataFrame(), "bacen_detail")}
            </section>

            <section class="section">
                <h2 class="section-title">Validações Cadastrais Brasileiras</h2>
                <p class="lead">Validações semânticas de CPF, CNPJ, UF, município + UF, código IBGE e formato de CEP. Documentos são mascarados nas amostras de erro.</p>
                {df_to_html(semantic_df if semantic_df is not None else pd.DataFrame(), "semantic")}
            </section>

            <section class="section">
                <h2 class="section-title">Amostras de Inconsistências Cadastrais</h2>
                <p class="lead">Amostra limitada de registros inválidos para apoio à correção, sem exposição integral de CPF ou CNPJ.</p>
                {df_to_html(semantic_invalid_df if semantic_invalid_df is not None else pd.DataFrame(), "semantic_invalid")}
            </section>

            <section class="section">
                <h2 class="section-title">Colunas com Maior Atenção</h2>
                {df_to_html(attention_df, "attention")}
            </section>

            <section class="section">
                <h2 class="section-title">Detalhe Técnico</h2>
                <p class="lead">Nesta seção, as notas técnicas originais de 0 a 10 são preservadas para rastreabilidade.</p>
                {df_to_html(detail_df, "technical", max_rows=500)}
            </section>

            <section class="section">
                <h2 class="section-title">Critérios para Interpretação da Aderência de Qualidade</h2>
                <p class="lead">As faixas e definições abaixo padronizam a leitura executiva da aderência, das classificações e dos tipos de avaliação utilizados nas dimensões BACEN.</p>
                {adherence_range_html()}
                {bacen_evaluation_glossary_html()}
            </section>
        </div>

        <div class="footer">MJV Data Quality • Relatório Executivo</div>
    </div>
</body>
</html>
"""
    path.write_text(html, encoding="utf-8")
    print(f"[OK] HTML : {path}")


# =========================================================
# Processo principal
# =========================================================
def run(duckdb_path, schema, outdir, run_id=None, logo=None):
    outdir_path = ensure_dir(outdir)
    con = duckdb.connect(duckdb_path)

    effective_run_id = run_id or get_latest_run_id(con, schema)
    if not effective_run_id:
        raise RuntimeError("Nenhum run_id encontrado.")

    summary_df = load_summary(con, schema, effective_run_id)
    detail_df = load_detail(con, schema, effective_run_id)

    if summary_df.empty and not detail_df.empty:
        summary_df = derive_summary_from_detail(detail_df, effective_run_id)

    input_dir = Path(duckdb_path).parent / "input"
    summary_df = enrich_summary_with_record_type_metrics(summary_df, input_dir)

    bacen_df = load_bacen_dimensions(con, schema, effective_run_id)
    bacen_detail_df = load_bacen_dimension_detail(con, schema, effective_run_id)
    semantic_df = load_semantic_results(con, schema, effective_run_id)
    semantic_invalid_df = load_semantic_invalid_samples(con, schema, effective_run_id)
    dim_df = bacen_df if not bacen_df.empty else build_dimensions(summary_df, effective_run_id)
    dim_df = apply_na_to_unmeasured_dimensions(dim_df, bacen_detail_df)
    bacen_catalog_df = build_bacen_catalog(dim_df, bacen_detail_df)
    history_df = load_history(con, schema)
    kpis = build_kpis(summary_df)
    top_df, bottom_df = build_top_bottom(summary_df)
    attention_df = build_attention(detail_df)

    genai_result = generate_genai_analysis(
        run_id=effective_run_id,
        kpis=kpis,
        summary_df=summary_df,
        dim_df=dim_df,
        bacen_detail_df=bacen_detail_df,
        semantic_df=semantic_df,
    )
    genai_json_path = outdir_path / "dq_genai_analysis_current.json"
    genai_json_path.write_text(
        json.dumps(genai_result, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Arquivos premium principais
    premium_xlsx = outdir_path / f"dq_report_premium_mjv_v2_{timestamp}.xlsx"
    premium_html = outdir_path / f"dq_report_premium_mjv_v2_{timestamp}.html"
    export_excel(premium_xlsx, kpis, summary_df, dim_df, detail_df, top_df, bottom_df, attention_df, history_df, bacen_detail_df, bacen_catalog_df, semantic_df, semantic_invalid_df, genai_result)
    export_html(premium_html, effective_run_id, kpis, summary_df, dim_df, detail_df, top_df, bottom_df, attention_df, history_df, bacen_detail_df, bacen_catalog_df, semantic_df, semantic_invalid_df, genai_result)

    # Arquivos executivos compatíveis com pipeline atual
    exec_xlsx = outdir_path / f"dq_executive_report_v2_{timestamp}.xlsx"
    exec_html = outdir_path / f"dq_executive_report_v2_{timestamp}.html"
    export_excel(exec_xlsx, kpis, summary_df, dim_df, detail_df, top_df, bottom_df, attention_df, history_df, bacen_detail_df, bacen_catalog_df, semantic_df, semantic_invalid_df, genai_result)
    export_html(exec_html, effective_run_id, kpis, summary_df, dim_df, detail_df, top_df, bottom_df, attention_df, history_df, bacen_detail_df, bacen_catalog_df, semantic_df, semantic_invalid_df, genai_result)

    export_support_files(outdir_path, effective_run_id, detail_df, dim_df, history_df, attention_df, bacen_detail_df)

    print(f"[OK] Excel: {exec_xlsx}")
    print(f"[OK] HTML : {exec_html}")
    print(f"[OK] Recomendações IA: {outdir_path / 'dq_ai_recommendations_current.csv'}")
    print(f"[OK] Análise GenAI: {genai_json_path}")
    print(f"[OK] Detalhe atual: {outdir_path / 'dq_current_detail.csv'}")
    print(f"[OK] Dimensões atuais: {outdir_path / 'dq_dimension_scores_current.csv'}")
    if "score_bacen" in dim_df.columns:
        print(f"[OK] Resumo BACEN: {outdir_path / 'dq_bacen_summary_current.csv'}")
    if not bacen_detail_df.empty:
        print(f"[OK] Detalhe BACEN: {outdir_path / 'dq_bacen_dimension_detail_current.csv'}")

    if not semantic_df.empty:
        semantic_path = outdir_path / "dq_semantic_validation_current.csv"
        semantic_df.to_csv(semantic_path, index=False, encoding="utf-8-sig")
        print(f"[OK] Validações Brasil: {semantic_path}")
    if not semantic_invalid_df.empty:
        invalid_path = outdir_path / "dq_semantic_invalid_samples_current.csv"
        semantic_invalid_df.to_csv(invalid_path, index=False, encoding="utf-8-sig")
        print(f"[OK] Amostras inválidas: {invalid_path}")

    con.close()


def load_semantic_results(con, schema, run_id):
    try:
        if not table_exists(con, schema, "dq_semantic_validation_results"):
            return pd.DataFrame()
        return fetchdf_safe(con, f"SELECT * FROM {schema}.dq_semantic_validation_results WHERE run_id = ? ORDER BY object_name, validation_code", [run_id])
    except Exception:
        return pd.DataFrame()


def load_semantic_invalid_samples(con, schema, run_id):
    try:
        if not table_exists(con, schema, "dq_semantic_invalid_samples"):
            return pd.DataFrame()
        return fetchdf_safe(con, f"SELECT * FROM {schema}.dq_semantic_invalid_samples WHERE run_id = ? ORDER BY object_name, validation_code, row_number", [run_id])
    except Exception:
        return pd.DataFrame()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--duckdb", required=True)
    parser.add_argument("--schema", default="stg")
    parser.add_argument("--outdir", default="./output")
    parser.add_argument("--run_id", default=None)
    parser.add_argument("--logo", default=None)
    args = parser.parse_args()

    run(
        duckdb_path=args.duckdb,
        schema=args.schema,
        outdir=args.outdir,
        run_id=args.run_id,
        logo=args.logo,
    )


if __name__ == "__main__":
    main()
